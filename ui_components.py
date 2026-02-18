import os
import tkinter as tk
from tkinter import ttk
from tkinter import ttk, filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from pathlib import Path
import json  
import logging
import datetime 

class BalancesTab:
    def __init__(self, parent, logger):
        self.parent = parent
        self.logger = logger
        self.setup_ui()
        self.rates = {}  # Store current rates
        self.update_callback = None  # Add callback storage
   
    def show_info_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showinfo(title, message, parent=self.root)
    
    def show_error_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showerror(title, message, parent=self.root)
    
    def show_warning_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showwarning(title, message, parent=self.root)

    def setup_ui(self):
        # Create main frame
        self.frame = ttk.Frame(self.parent)
        self.frame.grid(sticky='nsew')
        
        # Create left container for treeview and total
        left_container = ttk.Frame(self.frame)
        left_container.grid(row=0, column=0, sticky='nsew', padx=(0, 10))  # Add padding to the right
    
        # Create Treeview for balances
        columns = ('Währung', 'zur Verfügung', 'Reserviert', 'Total', 'EUR Equivalent')
        self.balances_tree = ttk.Treeview(
            left_container, 
            columns=columns,
            show='headings',
            height=12
        )
        
        # Set column headings and widths
        column_widths = {
            'Währung': 100,
            'zur Verfügung': 150,
            'Reserviert': 150,
            'Total': 150,
            'EUR Equivalent': 150
        }
        
        for col in columns:
            self.balances_tree.heading(col, text=col)
            self.balances_tree.column(col, width=column_widths[col], anchor='center')
    
        # Add scrollbar
        scrollbar = ttk.Scrollbar(
            left_container,
            orient=tk.VERTICAL,
            command=self.balances_tree.yview
        )
        
        self.balances_tree.configure(yscrollcommand=scrollbar.set)
    
        # Grid layout for treeview and scrollbar
        self.balances_tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
    
        # Add total EUR value label
        self.total_label = ttk.Label(left_container, text="Total Portfolio Stand: 0.00 EUR")
        self.total_label.grid(row=1, column=0, pady=10, sticky='e')
    
        # Create right container for button
        right_container = ttk.Frame(self.frame)
        right_container.grid(row=0, column=1, sticky='ns')  # Changed sticky to 'ns'
        
        # Add update button with explicit width
        self.update_button = ttk.Button(
            right_container,
            text="Kontostände Aktualisieren",
            command=self.update_pressed,
            width=25  # Added explicit width
        )
        self.update_button.grid(row=0, column=0, pady=(80, 0))  # Adjust vertical position
    
        # Configure grid weights
        self.frame.columnconfigure(0, weight=1)  # Make left container expandable
        self.frame.columnconfigure(1, weight=0)  # Keep right container fixed width
        left_container.columnconfigure(0, weight=1)  # Make treeview expandable
        
    def set_update_callback(self, callback):
        """Set the callback function for the update button"""
        self.update_callback = callback

    def update_pressed(self):
        """Handle update button press"""
        if self.update_callback:
            try:
                self.update_button.configure(state='disabled')  # Disable button during update
                self.update_callback()  # Call the callback function
                self.logger.info("Balance update requested")
            except Exception as e:
                self.logger.error(f"Error during balance update: {str(e)}")
            finally:
                self.update_button.configure(state='normal')  # Re-enable button
        else:
            self.logger.warning("Update callback not set")

    def update_rates(self, rates_data):
        """Update stored rates"""
        try:
            self.rates = rates_data
            self.logger.debug(f"Updated rates: {json.dumps(rates_data, indent=2)}")
        except Exception as e:
            self.logger.error(f"Error updating rates: {str(e)}")
    
    def calculate_eur_equivalent(self, currency, amount):
        """Calculate EUR equivalent for given currency and amount"""
        try:
            if currency.lower() == 'eur':
                return float(amount)
    
            rate_key = f"{currency.lower()}eur"
            if rate_key in self.rates:
                rate = self.rates[rate_key].get('rate', {}).get('rate_weighted')
                if rate and rate != 'N/A':
                    return float(amount) * float(rate)
                return 0
            return 0
        except Exception as e:
            self.logger.error(f"Error calculating EUR equivalent for {currency}: {str(e)}")
            return 0
          
    def update_balances(self, data):
        # Clear existing items
        for item in self.balances_tree.get_children():
            self.balances_tree.delete(item)

        try:
            total_eur_value = 0
            balances = data.get('data', {}).get('balances', {})
            
            for currency, balance in balances.items():
                decimals = 2 if currency.lower() == 'eur' else 8
                
                available = float(balance.get('available_amount', 0))
                reserved = float(balance.get('reserved_amount', 0))
                total = float(balance.get('total_amount', 0))
                
                # Calculate EUR equivalent
                eur_equivalent = self.calculate_eur_equivalent(currency, total)
                total_eur_value += eur_equivalent

                self.balances_tree.insert('', 'end', values=(
                    currency.upper(),
                    f"{available:.{decimals}f}",
                    f"{reserved:.{decimals}f}",
                    f"{total:.{decimals}f}",
                    f"{eur_equivalent:,.2f} €"
                ))
                
                self.logger.debug(f"Processed {currency}: {total} = {eur_equivalent} EUR")

            # Update total portfolio value
            self.total_label.configure(text=f"Total Portfolio Value: {total_eur_value:,.2f} €")
            self.logger.info(f"Total portfolio value: {total_eur_value:,.2f} EUR")
            
        except Exception as e:
            self.logger.error(f"Error updating balances: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
    
import tkinter as tk
from tkinter import ttk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import StringVar
from datetime import datetime
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt

class RatesTab:
    def __init__(self, parent, logger, db_manager):
        self.parent = parent
        self.logger = logger
        self.db_manager = db_manager
        self.rates_data = {}  # Add this to store rates
        self.analysis_data = {}  # Add this to store analysis data
        self.gauges = {}  # Store gauges for each trading pair
        self.selected_interval = StringVar(value='Halber Tag')  # Default interval
        self.interval_mapping = {
            'Halber Tag': '30m',
            'Tag': '1h',
            'Woche': '8h',
            'Monat': '24h'
        }
        self.setup_ui()
        self.update_analysis_data()  # Initial data load
        self.schedule_updates()  # Schedule periodic updates

    def show_info_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showinfo(title, message, parent=self.root)
    
    def show_error_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showerror(title, message, parent=self.root)
    
    def show_warning_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showwarning(title, message, parent=self.root)

    def setup_ui(self):
        # Create main frame
        self.frame = ttk.Frame(self.parent)
        self.frame.grid(sticky='nsew')
        
        # Create left container for treeview
        left_container = ttk.Frame(self.frame)
        left_container.grid(row=0, column=0, sticky='nsew', padx=(0, 10))

        # Create Treeview for rates
        columns = ('Handelspaar', 'gewichtiger Kurs aktuell', 'gewichtiger Kurs 3h', 'gewichtiger Kurs 12h')
        self.rates_tree = ttk.Treeview(
            left_container, 
            columns=columns,
            show='headings',
            height=10
        )
        
        # Set column headings and widths
        column_widths = {
            'Handelspaar': 150,
            'gewichtiger Kurs aktuell': 150,
            'gewichtiger Kurs 3h': 150,
            'gewichtiger Kurs 12h': 150
        }
        
        for col in columns:
            self.rates_tree.heading(col, text=col, anchor='center')
            self.rates_tree.column(col, width=column_widths[col], anchor='center')

        # Add scrollbar
        scrollbar = ttk.Scrollbar(
            left_container,
            orient=tk.VERTICAL,
            command=self.rates_tree.yview
        )
        
        self.rates_tree.configure(yscrollcommand=scrollbar.set)
        self.rates_tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        # Create right container for button
        button_container = ttk.Frame(self.frame)
        button_container.grid(row=0, column=1, sticky='ns')
        
        # Add update button with explicit width
        self.update_button = ttk.Button(
            button_container,
            text="Kurse Aktualisieren",
            command=self.update_pressed,
            width=25
        )
        self.update_button.grid(row=0, column=0, pady=(80, 0))

        # Add interval selection combobox
        self.interval_combobox = ttk.Combobox(
            button_container,
            textvariable=self.selected_interval,
            values=list(self.interval_mapping.keys()),
            state='readonly',
            width=22  # Adjust width to match button
        )
        self.interval_combobox.grid(row=1, column=0, pady=(10, 0))  # Position below the update button

        # Bind the combobox selection event to update the analysis data
        self.interval_combobox.bind('<<ComboboxSelected>>', self.on_interval_changed)

        # Create right container for analysis data
        analysis_container = ttk.Frame(self.frame)
        analysis_container.grid(row=1, column=0, columnspan=2, sticky='nsew')

        # Create blocks for each trading pair
        self.blocks = {}
        trading_pairs = ['btceur', 'bcheur', 'etheur', 'soleur', 'xrpeur', 'ltceur', 'dogeeur', 'btgeur', 'trxeur', 'usdceur']
        for i, pair in enumerate(trading_pairs):
            block = ttk.LabelFrame(analysis_container, text=pair.upper(), padding="10")
            block.grid(row=i//5, column=i%5, padx=10, pady=10, sticky='nsew')
            self.blocks[pair] = block

            # Add labels for analysis data
            ttk.Label(block, text="RSI:").grid(row=0, column=0, sticky='w')
            ttk.Label(block, text="BB SMA:").grid(row=1, column=0, sticky='w')
            ttk.Label(block, text="Stochastic:").grid(row=2, column=0, sticky='w')
            ttk.Label(block, text="Score:").grid(row=3, column=0, sticky='w')

            self.analysis_data[pair] = {
                'rsi_value': ttk.Label(block, text="N/A"),
                'bb_sma': ttk.Label(block, text="N/A"),
                'stochastic_value': ttk.Label(block, text="N/A"),
                'score': ttk.Label(block, text="N/A")
            }
            self.analysis_data[pair]['rsi_value'].grid(row=0, column=1, sticky='e')
            self.analysis_data[pair]['bb_sma'].grid(row=1, column=1, sticky='e')
            self.analysis_data[pair]['stochastic_value'].grid(row=2, column=1, sticky='e')
            self.analysis_data[pair]['score'].grid(row=3, column=1, sticky='e')

            # Add gauge for score
            self.add_gauge(block, pair)

        # Configure grid weights
        self.frame.columnconfigure(0, weight=1)
        self.frame.columnconfigure(1, weight=1)
        left_container.columnconfigure(0, weight=1)
        analysis_container.columnconfigure(0, weight=1)
        for i in range(5):
            analysis_container.columnconfigure(i, weight=1)

    def add_gauge(self, parent, pair):
        """Add a gauge to the given parent frame"""
        fig, ax = plt.subplots(figsize=(2, 1))  # Adjust size to make it a half-circle
    
        ax.axis('off')
        self.gauges[pair] = ax

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.get_tk_widget().grid(row=4, column=0, columnspan=2, pady=10)
        self.update_gauge(pair, 0)  # Initialize with 0 value

    def update_gauge(self, pair, value):
        """Update the gauge for the given pair with the specified value"""
        ax = self.gauges[pair]
        ax.clear()
        ax.axis('off')
    
        # Create the gauge
        min_value, max_value = -3, 3
        start_angle, end_angle = 180, 0
        theta = np.linspace(start_angle, end_angle, 100)
        radius = 1
    
        # Draw the gauge background
        ax.plot(radius * np.cos(np.radians(theta)), radius * np.sin(np.radians(theta)), color='lightgrey', lw=10)
    
        # Draw the colored ranges
        range_angles = np.linspace(start_angle, end_angle, 4)
        colors = ['red', 'yellow', 'green', 'yellow']
        for i in range(len(range_angles) - 1):
            ax.plot(radius * np.cos(np.radians(np.linspace(range_angles[i], range_angles[i + 1], 100))),
                    radius * np.sin(np.radians(np.linspace(range_angles[i], range_angles[i + 1], 100))), color=colors[i], lw=10)
    
        # Draw the needle
        needle_angle = start_angle - (value - min_value) / (max_value - min_value) * (start_angle - end_angle)
        ax.plot([0, radius * np.cos(np.radians(needle_angle))], [0, radius * np.sin(np.radians(needle_angle))], color='black', lw=2)
    
        # Draw the center circle
        ax.plot(0, 0, 'o', color='black')
    
        # Draw the value text
        ax.text(0, -0.15, f'{value:.2f}', ha='center', va='center', fontsize=8)
    
        # Redraw the canvas
        ax.figure.canvas.draw()

    def update_analysis_data(self):
        """Update analysis data for all trading pairs"""
        if not self.db_manager:
            self.logger.warning("Database manager is not initialized. Skipping analysis data retrieval.")
            return

        interval_display = self.selected_interval.get()  # Get the selected interval display value
        interval = self.interval_mapping[interval_display]  # Map to the actual interval value
        trading_pairs = ['btceur', 'bcheur', 'etheur', 'soleur', 'xrpeur', 'ltceur', 'dogeeur', 'btgeur', 'trxeur', 'usdceur']
        for pair in trading_pairs:
            data = self.db_manager.get_analysis_data(pair, interval)  # Pass the interval to the database manager
            if data:
                self.analysis_data[pair]['rsi_value'].configure(text=f"{data['rsi_value']:.2f}" if data['rsi_value'] is not None else "N/A")
                self.analysis_data[pair]['bb_sma'].configure(text=f"{data['bb_sma']:.2f}" if data['bb_sma'] is not None else "N/A")
                self.analysis_data[pair]['stochastic_value'].configure(text=f"{data['stochastic_value']:.2f}" if data['stochastic_value'] is not None else "N/A")
                self.analysis_data[pair]['score'].configure(text=f"{data['score']:.2f}" if data['score'] is not None else "N/A")
                self.update_gauge(pair, data['score'] if data['score'] is not None else 0)
            else:
                self.analysis_data[pair]['rsi_value'].configure(text="N/A")
                self.analysis_data[pair]['bb_sma'].configure(text="N/A")
                self.analysis_data[pair]['stochastic_value'].configure(text="N/A")
                self.analysis_data[pair]['score'].configure(text="N/A")
                self.update_gauge(pair, 0)

    def on_interval_changed(self, event):
        """Handle interval change event"""
        self.update_analysis_data()

    def schedule_updates(self):
        """Schedule periodic updates for analysis data"""
        self.update_analysis_data()
        update_interval_minutes = 5  # Set the update interval in minutes
        self.parent.after(update_interval_minutes * 60 * 1000, self.schedule_updates)  # Convert minutes to milliseconds
    
    def set_update_callback(self, callback):
        """Set the callback function for the update button"""
        self.update_callback = callback

    def update_pressed(self):
        """Handle update button press"""
        if self.update_callback:
            try:
                self.update_button.configure(state='disabled')  # Disable button during update
                self.update_callback()  # Call the callback function
                self.logger.info("Rates update requested")
            except Exception as e:
                self.logger.error(f"Error during rates update: {str(e)}")
            finally:
                self.update_button.configure(state='normal')  # Re-enable button
        else:
            self.logger.warning("Update callback not set")

    def clear_rates(self):
        """Clear all rates from the tree"""
        for item in self.rates_tree.get_children():
            self.rates_tree.delete(item)

    def update_rates(self, rates_data):
        try:
            # Get the rate data from the response
            rate_data = rates_data.get('rate', {})
            trading_pair = rates_data.get('trading_pair', '').upper()
            
            if not rate_data or not trading_pair:
                return
    
            # Get the weighted rates
            rate = rate_data.get('rate_weighted', 'N/A')
            rate_3h = rate_data.get('rate_weighted_3h', 'N/A')
            rate_12h = rate_data.get('rate_weighted_12h', 'N/A')
    
            # Insert the rates
            self.rates_tree.insert('', 'end', values=(
                trading_pair,
                self.format_price(rate),
                self.format_price(rate_3h),
                self.format_price(rate_12h)
            ))
    
        except Exception as e:
            self.logger.error(f"Error updating rates: {str(e)}")

    def format_price(self, price):
        """Format price value"""
        try:
            if price == 'N/A':
                return price
            return f"{float(price):.4f}"
        except (ValueError, TypeError):
            return 'N/A'
        
    def stop_auto_updates(self):
        """Stop automatic updates"""
        if hasattr(self, 'after_id'):
            self.parent.after_cancel(self.after_id)
        
import tkinter as tk
from tkinter import ttk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import StringVar, IntVar, BooleanVar
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import threading
import time

class OrderbookTab:
    def __init__(self, parent, logger, db_manager=None, trading_tab=None, api_client=None):  # Added db_manager parameter with default None
        self.parent = parent
        self.logger = logger
        self.db_manager = db_manager
        self.selected_pair = tk.StringVar(value="Bitcoin (BTC/EUR)")  # Default value
        self.trading_tab = trading_tab  # Store reference to trading tab
        self.api_client = api_client
    
        # Add variables to track previous counts
        self.previous_ask_count = 0
        self.previous_bid_count = 0
        self.update_interval = 500  # Update every 1/2 second
        if self.db_manager:  # Only start auto-updates if db_manager exists
            self.start_auto_updates()

        self.current_pair = None  # Add this to track the current pair
        self.previous_orders = {'asks': set(), 'bids': set()}  # Store previous orders
        self.flash_duration = 500  # Duration of flash effect in milliseconds

        self.setup_ui()
        self.create_context_menu()
        self.order_ids = {}

    def show_info_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showinfo(title, message, parent=self.root)
    
    def show_error_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showerror(title, message, parent=self.root)
    
    def show_warning_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showwarning(title, message, parent=self.root)
        
    def setup_ui(self):
        # Create main frame
        main_frame = ttk.Frame(self.parent)
        main_frame.grid(row=0, column=0, sticky='nsew')
    
        # Create currency selector frame
        selector_frame = ttk.LabelFrame(main_frame, text="Handelspaare KRYPTO / EURO", padding="5")
        selector_frame.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='ew')
    
        # Add currency selector
        ttk.Label(selector_frame, text="Handelspaar auswählen:").grid(row=0, column=0, padx=5, pady=5)
        
        # Create combobox with trading pairs
        self.pairs = [
            ("Bitcoin (BTC/EUR)", "btceur"),
            ("Bitcoin Cash (BCH/EUR)", "bcheur"),
            ("Ethereum (ETH/EUR)", "etheur"),
            ("Solana (SOL/EUR)", "soleur"),
            ("Ripple (XRP/EUR)", "xrpeur"),
            ("Litecoin (LTC/EUR)", "ltceur"),
            ("Dogecoin (DOGE/EUR)", "dogeeur"),
            ("Bitcoin Gold (BTG/EUR)", "btgeur"),
            ("TRON (TRX/EUR)", "trxeur"),
            ("USDC (USDC/EUR)", "usdceur")
        ]
        
        self.pair_combobox = ttk.Combobox(
            selector_frame, 
            textvariable=self.selected_pair,
            values=[pair[0] for pair in self.pairs],
            state='readonly',
            width=30
        )
        self.pair_combobox.grid(row=0, column=1, padx=5, pady=5)
        self.pair_combobox.set("Bitcoin (BTC/EUR)")  # Set default value
        
        # Bind the combobox selection event
        self.pair_combobox.bind('<<ComboboxSelected>>', self.on_pair_changed)
       
        # Store the mapping of display names to API values
        self.pair_mapping = {pair[0]: pair[1] for pair in self.pairs}
    
        # Create frames for asks and bids
        self.asks_frame = ttk.LabelFrame(main_frame, text="Verkäufe (Sell Orders)")
        self.bids_frame = ttk.LabelFrame(main_frame, text="Kaufaufträge (Buy Orders)")
        
        # Position frames (asks left, bids right)
        self.asks_frame.grid(row=1, column=0, padx=5, pady=5, sticky='nsew')
        self.bids_frame.grid(row=1, column=1, padx=5, pady=5, sticky='nsew')
    
        # Create Treeviews
        columns = ('Kurs', 'Volumen', 'Total Preis')
        
        self.asks_tree = self.create_treeview(self.asks_frame, columns)
        self.bids_tree = self.create_treeview(self.bids_frame, columns)

        # Bind the context menu to the Treeviews
        self.asks_tree.bind('<Button-3>', self.show_context_menu)
        self.bids_tree.bind('<Button-3>', self.show_context_menu)
    
        # Configure grid weights
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        selector_frame.columnconfigure(1, weight=1)

    def create_context_menu(self):
        self.context_menu = tk.Menu(self.parent, tearoff=0)
        self.context_menu.add_command(label="Orderdetails anzeigen", command=self.show_order_details)
    
        # Create "Order annehmen" submenu
        self.order_submenu = tk.Menu(self.context_menu, tearoff=0)
        self.context_menu.add_cascade(label="Order annehmen", menu=self.order_submenu)
    
        # Create "Jetzt verkaufen" menu item
        self.order_submenu.add_command(label="Jetzt verkaufen", command=lambda: self.take_order('sell', payment_option=2))
        self.sell_now_item_index = self.order_submenu.index("end")
    
        # Create "Jetzt min verkaufen" menu item
        self.order_submenu.add_command(label="Jetzt min verkaufen", command=lambda: self.take_order('sell', payment_option=2, min_amount=True))
        self.sell_now_min_item_index = self.order_submenu.index("end")
    
        # Create "Sell Order" submenu for payment options
        self.sell_submenu = tk.Menu(self.order_submenu, tearoff=0)
        self.sell_submenu.add_command(label="Express-Kauf", command=lambda: self.take_order('buy', 1))
        self.sell_submenu.add_command(label="SEPA-Kauf", command=lambda: self.take_order('buy', 2))
        self.sell_submenu.add_command(label="SEPA Echtzeitüberweisung", command=lambda: self.take_order('buy', 3))
    
        # Create "Kaufen mit Zahlungsmethode" menu item
        self.order_submenu.add_cascade(label="Kaufen mit Zahlungsmethode", menu=self.sell_submenu)
        self.buy_with_payment_item_index = self.order_submenu.index("end")
    
        # Create "min Kaufen mit Zahlungsmethode" submenu for payment options
        self.min_buy_submenu = tk.Menu(self.order_submenu, tearoff=0)
        self.min_buy_submenu.add_command(label="Express-Kauf", command=lambda: self.take_order('buy', 1, min_amount=True))
        self.min_buy_submenu.add_command(label="SEPA-Kauf", command=lambda: self.take_order('buy', 2, min_amount=True))
        self.min_buy_submenu.add_command(label="SEPA Echtzeitüberweisung", command=lambda: self.take_order('buy', 3, min_amount=True))
    
        # Create "min Kaufen mit Zahlungsmethode" menu item
        self.order_submenu.add_cascade(label="min Kaufen mit Zahlungsmethode", menu=self.min_buy_submenu)
        self.buy_with_min_payment_item_index = self.order_submenu.index("end")
    
    def show_context_menu(self, event):
        tree = event.widget
        item = tree.identify_row(event.y)
        if item:
            self.selected_order = tree.item(item, 'values')
            self.selected_order_id = self.order_ids.get(item)  # Retrieve the hidden order_id
    
            # Determine if the selected order is a buy or sell order
            if tree == self.asks_tree:
                # It's a sell order
                self.order_submenu.entryconfig(self.sell_now_item_index, state='disabled')
                self.order_submenu.entryconfig(self.sell_now_min_item_index, state='disabled')
                self.order_submenu.entryconfig(self.buy_with_payment_item_index, state='normal')
                self.order_submenu.entryconfig(self.buy_with_min_payment_item_index, state='normal' if len(self.selected_order) > 4 and self.selected_order[4] else 'disabled')
            elif tree == self.bids_tree:
                # It's a buy order
                self.order_submenu.entryconfig(self.sell_now_item_index, state='normal')
                self.order_submenu.entryconfig(self.sell_now_min_item_index, state='normal' if len(self.selected_order) > 4 and self.selected_order[4] else 'disabled')
                self.order_submenu.entryconfig(self.buy_with_payment_item_index, state='disabled')
                self.order_submenu.entryconfig(self.buy_with_min_payment_item_index, state='disabled')
    
            self.context_menu.post(event.x_root, event.y_root)

    def take_order(self, order_type, payment_option=None, min_amount=False):
        if self.selected_order and len(self.selected_order) >= 4:
            price = float(self.selected_order[0].replace(' €', '').replace(',', ''))
            amount = float(self.selected_order[1].split()[0])
            trading_pair = self.get_selected_pair_api_value()
            order_id = self.selected_order[3]  # Assuming order_id is the fourth value in selected_order
    
            if min_amount and len(self.selected_order) > 4 and self.selected_order[4]:
                try:
                    amount = float(self.selected_order[4].split()[0])
                except ValueError:
                    self.logger.error(f"Invalid min_amount value: {self.selected_order[4]}")
                    messagebox.showerror("Fehler", "Ungültiger Wert für min_amount.")
                    return
    
            # Use the correct order_type directly in the API request
            try:
                if order_type == 'buy':
                    response = self.api_client.take_buy_order(trading_pair, order_id, amount, payment_option)
                else:
                    response = self.api_client.take_order(trading_pair, order_id, order_type, amount, payment_option)
    
                messagebox.showinfo("Erfolg", f"Order erfolgreich {order_type}!")
            except Exception as e:
                self.logger.error(f"Fehler beim {order_type} der Order: {str(e)}")
                messagebox.showerror("Fehler", f"Fehler beim {order_type} der Order: {str(e)}")
        else:
            self.logger.error("Selected order does not contain enough values")
            messagebox.showerror("Fehler", "Die ausgewählte Order enthält nicht genügend Werte.")

    def show_order_details(self):
        if self.selected_order:
            price = float(self.selected_order[0].replace(' €', '').replace(',', ''))
            amount = float(self.selected_order[1].split()[0])
            total_price = float(self.selected_order[2].replace(' €', '').replace(',', ''))
    
            # Ensure min_amount is correctly extracted
            min_amount = None
            if len(self.selected_order) > 4 and self.selected_order[4]:
                try:
                    min_amount = float(self.selected_order[4].split()[0])
                except ValueError:
                    min_amount = None
    
            details = (
                f"Order ID: {self.selected_order_id}\n"
                f"Preis: {self.selected_order[0]}\n"
                f"Volumen: {self.selected_order[1]}\n"
                f"Total Preis: {self.selected_order[2]}"
            )
    
            if min_amount and min_amount != amount:
                min_total_price = min_amount * price
                details += (
                    f"\nmin Volumen: {min_amount:.8f}\n"
                    f"min Total Preis: {min_total_price:,.2f} €"
                )
    
            messagebox.showinfo("Orderdetails", details)

    def create_treeview(self, parent, columns):
        """Create a treeview with scrollbar"""
        # Create frame to hold treeview and scrollbar
        frame = ttk.Frame(parent)
        frame.grid(sticky='nsew')
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        # Create treeview
        tree = ttk.Treeview(frame, columns=columns, show='headings', height=28)
        
        # Set column headings
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=185, anchor='center')

        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        # Grid layout
        tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        return tree

    def get_selected_pair_api_value(self):
        """Get the API value for the selected trading pair"""
        display_name = self.selected_pair.get()
        return self.pair_mapping.get(display_name, 'btceur')

    def get_selected_currency_symbol(self):
        """Get the currency symbol for the selected pair"""
        pair = self.get_selected_pair_api_value()
        return pair[:3].upper()  # First three characters uppercase

    def on_pair_changed(self, *args):
        """Handle trading pair change"""
        if self.db_manager:
            self.update_from_database()
 
    def update_from_database(self):
        """Update orderbook data from database"""
        if not self.db_manager:
            return
                
        try:
            pair = self.get_selected_pair_api_value()
            total_orders = self.db_manager.check_orders(pair)
            data = self.db_manager.get_orderbook(pair)
            self.update_orderbook(data)
        except Exception as e:
            self.logger.error(f"Error updating from database: {str(e)}")

    def flash_item(self, tree, item, color):
        """Apply flash effect to a tree item"""
        original_tags = tree.item(item, 'tags')
        tree.tag_configure('flash_green', background='#90EE90')  # Light green
        tree.tag_configure('flash_red', background='#FFB6C1')    # Light red
        
        if color == 'green':
            tree.item(item, tags=('flash_green',))
        elif color == 'red':
            tree.item(item, tags=('flash_red',))
        
        # Schedule removal of flash effect
        tree.after(self.flash_duration, lambda: tree.item(item, tags=original_tags))
    
    def update_orderbook(self, data):
        """Update the orderbook display"""
        try:
            # Get the orderbook data
            orderbook = data.get('orders', {})
            currency_symbol = self.get_selected_currency_symbol()
            current_pair = self.get_selected_pair_api_value()
    
            # Create sets of current orders
            current_asks = {(float(ask[0]), float(ask[1])) for ask in orderbook.get('asks', []) if len(ask) >= 2}
            current_bids = {(float(bid[0]), float(bid[1])) for bid in orderbook.get('bids', []) if len(bid) >= 2}
    
            # Find deleted orders (present in previous but not in current)
            deleted_asks = self.previous_orders['asks'] - current_asks
            deleted_bids = self.previous_orders['bids'] - current_bids
    
            # Flash red for deleted orders
            for price, amount in deleted_asks:
                total = price * amount
                item = self.asks_tree.insert('', 'end', values=(
                    f"{price:,.2f} €",
                    f"{amount:.8f} {currency_symbol}",
                    f"{total:,.2f} €"
                ))
                self.flash_item(self.asks_tree, item, 'red')
                self.asks_tree.after(self.flash_duration, lambda item=item: self.asks_tree.delete(item))
    
            for price, amount in deleted_bids:
                total = price * amount
                item = self.bids_tree.insert('', 'end', values=(
                    f"{price:,.2f} €",
                    f"{amount:.8f} {currency_symbol}",
                    f"{total:,.2f} €"
                ))
                self.flash_item(self.bids_tree, item, 'red')
                self.bids_tree.after(self.flash_duration, lambda item=item: self.bids_tree.delete(item))
    
            # Clear existing items (except the ones being deleted)
            for item in self.asks_tree.get_children():
                if not self.asks_tree.item(item, 'tags'):  # Only delete items without flash tags
                    self.asks_tree.delete(item)
            for item in self.bids_tree.get_children():
                if not self.bids_tree.item(item, 'tags'):
                    self.bids_tree.delete(item)
    
            # Process new asks
            sorted_asks = sorted(current_asks, key=lambda x: x[0])  # Sort by price
            for ask in orderbook.get('asks', []):
                price, amount, min_amount, order_id = ask
                total = float(price) * float(amount)
                item = self.asks_tree.insert('', 'end', values=(
                    f"{float(price):,.2f} €",
                    f"{float(amount):.8f} {currency_symbol}",
                    f"{total:,.2f} €",
                    order_id,  # Add order_id as the fourth value
                    f"{float(min_amount):.8f}"  # Add min_amount as the fifth value
                ))
                self.order_ids[item] = order_id  # Store order_id in dictionary
                if (float(price), float(amount)) not in self.previous_orders['asks']:
                    self.flash_item(self.asks_tree, item, 'green')
    
            # Process new bids
            sorted_bids = sorted(current_bids, key=lambda x: x[0], reverse=True)  # Sort by price descending
            for bid in orderbook.get('bids', []):
                price, amount, min_amount, order_id = bid
                total = float(price) * float(amount)
                item = self.bids_tree.insert('', 'end', values=(
                    f"{float(price):,.2f} €",
                    f"{float(amount):,.8f} {currency_symbol}",
                    f"{total:,.2f} €",
                    order_id,  # Add order_id as the fourth value
                    f"{float(min_amount):.8f}"  # Add min_amount as the fifth value
                ))
                self.order_ids[item] = order_id  # Store order_id in dictionary
                if (float(price), float(amount)) not in self.previous_orders['bids']:
                    self.flash_item(self.bids_tree, item, 'green')
    
            # Only log if the trading pair has changed
            if current_pair != self.current_pair:
                self.logger.debug(f"Updated orderbook for {current_pair}: "
                                f"{len(current_asks)} asks, {len(current_bids)} bids")
                self.current_pair = current_pair
    
            # Update previous orders
            self.previous_orders['asks'] = current_asks
            self.previous_orders['bids'] = current_bids
    
        except Exception as e:
            self.logger.error(f"Error updating orderbook display: {str(e)}")
              
    def start_auto_updates(self):
        """Start automatic updates from database"""
        self.update_from_database()
        self.after_id = self.parent.after(self.update_interval, self.start_auto_updates)

    def stop_auto_updates(self):
        """Stop automatic updates"""
        if hasattr(self, 'after_id'):
            self.parent.after_cancel(self.after_id)

    def set_db_manager(self, db_manager):
        """Update the database manager and restart automatic updates"""
        self.stop_auto_updates()
        self.db_manager = db_manager
        if self.db_manager:
            self.start_auto_updates()

from tkinter import ttk, StringVar, IntVar, BooleanVar
from tkinter import messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from constants import TradingPairs, TradingConstants
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from constants import TradingPairs, TradingConstants  # Adjust the import path as needed
import threading
import time 

class TradingTab:
    def __init__(self, parent, logger, api_client=None):
        self.parent = parent
        self.logger = logger
        self.api_client = api_client
        
        # Create pair mapping from display names to API values
        self.pair_mapping = {
            display: api_name for api_name, display in TradingPairs.DISPLAY_NAMES.items()
        }

        # Initialize variables for trade settings
        self.selected_pair = StringVar(value=TradingPairs.DISPLAY_NAMES[TradingPairs.BTCEUR])
        self.trade_amount = StringVar(value="0.00")
        self.trade_price = StringVar(value="00.00")
        self.total_price = StringVar(value="0.00")
        self.trade_type = StringVar(value="buy")
        self.max_price = StringVar(value=TradingConstants.DEFAULT_MAX_PRICE)
        self.min_price = StringVar(value=TradingConstants.DEFAULT_MIN_PRICE)
        self.end_datetime = StringVar(value=self.get_default_end_datetime())
        self.min_trust_level = StringVar(value=TradingConstants.TrustLevel.BRONZE.value)
        self.payment_option = IntVar(value=TradingConstants.ORDERBOOK_PAYMENT_OPTION_SEPA_ONLY)
        self.sepa_option = IntVar(value=0)
        self.seat_of_bank = StringVar(value=TradingConstants.BankCountry.GERMANY.value)
        self.auto_trade = BooleanVar(value=False)

        # Define payment options mapping
        self.payment_options_map = {
            "Express & SEPA": 3,
            "SEPA-Only": 2,
            "Express-Only": 1
        }

        # Add SEPA options mapping
        self.sepa_options_map = {
            "SEPA Instant": 1,
            "SEPA Normal": 0
        }
    
        # Bind variables for automatic total calculation
        self.trade_amount.trace('w', self.calculate_total)
        self.trade_price.trace('w', self.calculate_total)

        # Initialize end_datetime with 2 days in future
        self.end_datetime = StringVar(value=self.get_default_end_datetime())

        # My Orders
        self.my_orders = set()  # Store order IDs of your orders
        self.orders_update_running = False
            
        self.setup_ui()

    def show_info_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showinfo(title, message, parent=self.root)
    
    def show_error_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showerror(title, message, parent=self.root)
    
    def show_warning_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showwarning(title, message, parent=self.root)

    def setup_ui(self):
        # Create main frame
        self.main_frame = ttk.Frame(self.parent)
        self.main_frame.pack(fill='both', expand=True)
    
        # Create controls frame
        self.controls_frame = ttk.LabelFrame(self.main_frame, text="Trade Einstellungen")
        self.controls_frame.pack(fill='x', padx=5, pady=(5, 0))
    
        # Create trading pair selection frame
        pair_frame = ttk.Frame(self.controls_frame)
        pair_frame.pack(fill='x', padx=1, pady=5)
    
        # Add trading pair label and dropdown
        ttk.Label(pair_frame, text="Handelspaar:").pack(side='left', padx=1)
        self.pair_dropdown = ttk.Combobox(
            pair_frame,
            textvariable=self.selected_pair,
            values=TradingPairs.get_all_display_names(),
            state="readonly",
            width=10
        )
        self.pair_dropdown.pack(side='left', padx=1)
    
        # Add trade amount entry
        ttk.Label(pair_frame, text="Menge:").pack(side='left', padx=1)
        self.amount_entry = ttk.Entry(pair_frame, textvariable=self.trade_amount, width=10)
        self.amount_entry.pack(side='left', padx=1)
    
        # Add trade price entry
        ttk.Label(pair_frame, text="Preis:").pack(side='left', padx=1)
        self.price_entry = ttk.Entry(pair_frame, textvariable=self.trade_price, width=10)
        self.price_entry.pack(side='left', padx=1)
    
        # Add total price display
        ttk.Label(pair_frame, text="Total (€):").pack(side='left', padx=1)
        self.total_label = ttk.Label(pair_frame, textvariable=self.total_price)
        self.total_label.pack(side='left', padx=1)
    
        # Add trade type selection with binding
        ttk.Label(pair_frame, text="Type:").pack(side='left', padx=1)
        self.type_dropdown = ttk.Combobox(
            pair_frame,
            textvariable=self.trade_type,
            values=["buy", "sell"],
            state="readonly",
            width=4
        )
        self.type_dropdown.pack(side='left', padx=1)
        self.type_dropdown.bind('<<ComboboxSelected>>', self.on_trade_type_changed)
    
        # Payment option selection frame
        self.payment_frame = ttk.Frame(pair_frame)
        self.payment_frame.pack(side='left')
    
        ttk.Label(self.payment_frame, text="Bezahl Option:").pack(side='left', padx=1)
        self.payment_option_combo = ttk.Combobox(
            self.payment_frame,
            values=list(self.payment_options_map.keys()),
            state="readonly",
            width=14
        )
        self.payment_option_combo.pack(side='left', padx=1)
        self.payment_option_combo.bind('<<ComboboxSelected>>', self.on_payment_option_changed)
    
        # SEPA option selection frame
        self.sepa_frame = ttk.Frame(pair_frame)
        self.sepa_frame.pack(side='left')
    
        ttk.Label(self.sepa_frame, text="SEPA Option:").pack(side='left', padx=1)
        self.sepa_option_combo = ttk.Combobox(
            self.sepa_frame,
            values=list(self.sepa_options_map.keys()),
            state="readonly",
            width=16
        )
        self.sepa_option_combo.pack(side='left', padx=1)
        self.sepa_option_combo.bind('<<ComboboxSelected>>', self.on_sepa_option_changed)
    
        # Add trust level selection
        ttk.Label(pair_frame, text="Trust Level:").pack(side='left', padx=1)
        self.trust_level_dropdown = ttk.Combobox(
            pair_frame,
            textvariable=self.min_trust_level,
            values=[level.value for level in TradingConstants.TrustLevel],
            state="readonly",
            width=8
        )
        self.trust_level_dropdown.pack(side='left', padx=1)

        # Bank country selection
        ttk.Label(pair_frame, text="Sitz der Bank:").pack(side='left', padx=1)
        self.bank_country_combo = ttk.Combobox(
            pair_frame,
            textvariable=self.seat_of_bank,
            values=[f"{code} {name}" for code, name in TradingConstants.BANK_COUNTRY_NAMES.items()],
            state="readonly",
            width=6
        )
        self.bank_country_combo.pack(side='left', padx=1)
    
        # Add end datetime entry with auto-update
        ttk.Label(pair_frame, text="Aktiv bis:").pack(side='left', padx=1)
        self.end_datetime_entry = ttk.Entry(pair_frame, textvariable=self.end_datetime, width=25)
        self.end_datetime_entry.pack(side='left', padx=1)
    
        # Add Create Order button
        self.trade_button = ttk.Button(
            self.controls_frame,
            text="Order einstellen",
            command=self.execute_trade
        )
        self.trade_button.pack(pady=10)
    
        # Initialize the UI state
        self.trade_type.set("buy")  # Set default to buy
        self.on_trade_type_changed(None)  # Update UI state
    
        # Start automatic updates
        self.update_end_datetime()

        # Add a separator
        ttk.Separator(self.main_frame, orient='horizontal').pack(fill='x', pady=10)

        # Create My Orders frame
        my_orders_frame = ttk.LabelFrame(self.main_frame, text="Meine aktiven Order")
        my_orders_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Create Treeview for orders
        columns = ('Order ID', 'Type', 'Handelspaar', 'Menge', 'Preis', 'beauftragt am')
        self.orders_tree = ttk.Treeview(
            my_orders_frame,
            columns=columns,
            show='headings',
            height=10,
            selectmode='browse'  # Single selection mode
        )

        # Bind right-click event
        self.orders_tree.bind('<Button-3>', self.show_context_menu)
        
        # Create context menu
        self.context_menu = tk.Menu(self.parent, tearoff=0)
        self.context_menu.add_command(label="Order Löschen", command=self.delete_selected_order)

        # Set column headings
        for col in columns:
            self.orders_tree.heading(col, text=col)
            self.orders_tree.column(col, width=100)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(
            my_orders_frame,
            orient='vertical',
            command=self.orders_tree.yview
        )
        self.orders_tree.configure(yscrollcommand=scrollbar.set)

        # Grid layout
        self.orders_tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        # Add Refresh button
        self.refresh_button = ttk.Button(
            my_orders_frame,
            text="Order Laden",
            command=self.start_orders_update
        )
        self.refresh_button.grid(row=1, column=0, pady=5)

        # Configure grid weights
        my_orders_frame.columnconfigure(0, weight=1)
        my_orders_frame.rowconfigure(0, weight=1)

        # Create context menu with more options
        self.context_menu = tk.Menu(self.parent, tearoff=0)
        self.context_menu.add_command(label="Order Löschen", command=self.delete_selected_order)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Order ID Kopieren", command=self.copy_order_id)
        self.context_menu.add_command(label="Order Laden", command=self.start_orders_update)

    def start_orders_update(self):
        """Start the orders update in a separate thread"""
        if not self.orders_update_running:
            self.orders_update_running = True
            self.refresh_button.configure(state='disabled')
            threading.Thread(target=self.update_my_orders, daemon=True).start()

    def update_my_orders(self):
        """Update the list of my orders"""
        try:
            response = self.api_client.get_orders()
            
            if 'orders' in response:
                # Clear existing items in the main thread
                self.parent.after(0, lambda: self.orders_tree.delete(*self.orders_tree.get_children()))
                
                # Update my_orders set
                self.my_orders.clear()
                
                # Process each order
                for order in response['orders']:
                    order_id = order.get('order_id')
                    if order_id:
                        self.my_orders.add(order_id)
                    
                    # Format the data
                    created_at = datetime.fromisoformat(order.get('created_at').replace('+01:00', '+0100')).strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Create values tuple
                    values = (
                        order_id,
                        order.get('type'),
                        order.get('trading_pair'),
                        f"{float(order.get('max_amount_currency_to_trade', 0)):.8f}",
                        f"{float(order.get('price', 0)):.2f} €",
                        created_at
                    )
                    
                    # Insert in the main thread
                    self.parent.after(0, lambda v=values: self.orders_tree.insert('', 'end', values=v))
    
                self.logger.info(f"Updated orders list: {len(self.my_orders)} active orders")
            else:
                self.logger.warning("No orders data in response")
    
        except Exception as e:
            self.logger.error(f"Error updating orders: {str(e)}")
        finally:
            self.orders_update_running = False
            self.parent.after(0, lambda: self.refresh_button.configure(state='normal'))
    
    def on_sepa_option_changed(self, _):
        """Handle SEPA option changes"""
        selected_text = self.sepa_option_combo.get()
        sepa_value = self.sepa_options_map.get(selected_text, 0)
        self.sepa_option.set(sepa_value)  # Update the IntVar with numeric value
        
    def on_trade_type_changed(self, _):
        """Handle trade type changes"""
        trade_type = self.trade_type.get()
        
        if trade_type == "sell":
            # Show payment options for sell orders
            self.payment_frame.pack(side='left')
            self.payment_option_combo.configure(state="readonly")
            
            # Set default to SEPA-Only
            self.payment_option_combo.set("SEPA-Only")
            self.payment_option.set(2)  # Set the IntVar to SEPA-Only (2)
            
            # Set default SEPA option to Instant
            self.sepa_frame.pack(side='left')
            self.sepa_option_combo.set("SEPA Instant")
            self.sepa_option.set(1)  # Set to SEPA Instant (1)
            
            # Update SEPA options state
            self.on_payment_option_changed(None)
        else:
            # Hide/disable payment options for buy orders
            self.payment_frame.pack_forget()
            self.sepa_frame.pack_forget()
            self.payment_option.set(0)
            self.sepa_option.set(0)
    
    def on_payment_option_changed(self, _):
        """Handle payment option changes"""
        if self.trade_type.get() == "sell":
            selected_text = self.payment_option_combo.get()
            payment_value = self.payment_options_map.get(selected_text, 0)
            self.payment_option.set(payment_value)  # Update the IntVar
            
            # Enable SEPA options only for SEPA-Only (2) or Express & SEPA (3)
            if payment_value in [2, 3]:
                self.sepa_frame.pack(side='left')
                self.sepa_option_combo.configure(state="readonly")
                # Set default to SEPA Instant if not already set
                if not self.sepa_option_combo.get():
                    self.sepa_option_combo.set("SEPA Instant")
                    self.sepa_option.set(1)
            else:
                self.sepa_frame.pack_forget()
                self.sepa_option.set(0)

    def execute_trade(self):
        """Execute a trade based on the current settings with a confirmation dialog."""
        # Sammle die aktuellen Einstellungen
        settings_summary = (
            f"Handelspaar: {self.selected_pair.get()}\n"
            f"Typ: {self.trade_type.get()}\n"
            f"Menge: {self.trade_amount.get()}\n"
            f"Preis: {self.trade_price.get()}\n"
            f"Enddatum: {self.end_datetime.get()}\n"
            f"Trust Level: {self.min_trust_level.get()}\n"
            f"Bezahloption: {self.payment_option_combo.get()}\n"
            f"SEPA Option: {self.sepa_option_combo.get()}\n"
            f"Sitz der Bank: {self.seat_of_bank.get()}"
        )
    
        # Zeige die Sicherheitsabfrage
        confirm = messagebox.askyesno(
            "Eingabewerte überprüfen",
            f"Bitte überprüfen Sie die folgenden Einstellungen:\n\n{settings_summary}\n\nSind diese korrekt?"
        )
    
        if not confirm:
            self.logger.info("Order-Erstellung abgebrochen durch den Benutzer.")
            return
    
        # Fortfahren mit der Order-Erstellung, wenn der Benutzer bestätigt hat
        if not self.api_client:
            self.logger.error("API client not configured")
            return
    
        try:
            # Get selected pair's API value
            selected_display_name = self.selected_pair.get()
            trading_pair = self.pair_mapping.get(selected_display_name)
    
            if not trading_pair:
                self.logger.error(f"Invalid trading pair: {selected_display_name}")
                return
    
            # Create base order parameters
            order_params = {
                "trading_pair": trading_pair.lower(),
                "type": self.trade_type.get(),
                "max_amount_currency_to_trade": float(self.trade_amount.get()),
                "price": float(self.trade_price.get()),
                "end_datetime": self.end_datetime.get() or None,
                "min_trust_level": self.min_trust_level.get(),
                "seat_of_bank": [self.seat_of_bank.get().split()[0]]  # Get country code
            }
    
            # Add payment options only for sell orders
            if self.trade_type.get() == "sell":
                payment_option = self.payment_option.get()
                if payment_option > 0:
                    order_params["payment_option"] = payment_option
    
                    # Add SEPA option only if applicable
                    if payment_option in [2, 3]:  # SEPA-Only or Express & SEPA
                        sepa_value = self.sepa_option.get()
                        if sepa_value is not None:
                            order_params["sepa_option"] = sepa_value
    
            # Disable the trade button to prevent multiple clicks
            self.trade_button.configure(state='disabled')
    
            def on_order_created(response):
                if 'order_id' in response:
                    order_id = response['order_id']
                    self.my_orders.add(order_id)  # Add to set of my orders
    
                    # Update the orders list in the background
                    self.start_orders_update()
    
                    self.logger.info(f"Trade erfolgreich eingestellt. Order ID: {order_id}")
                    messagebox.showinfo("Erfolg", f"Order erfolgreich erstellt\nOrder ID: {order_id}")
    
                    # Schedule a check for the order in the orderbook
                    self.schedule_order_check(order_id, order_params["type"])
                else:
                    self.logger.warning(f"Unexpected response format: {response}")
                    messagebox.showwarning("Warning", "Order Status unklar. Bitte Log prüfen.")
    
                # Re-enable the trade button
                self.trade_button.configure(state='normal')
    
            def on_order_failed(exception):
                self.logger.error(f"Error executing trade: {str(exception)}")
                messagebox.showerror("Error", f"Failed to create order: {str(exception)}")
    
                # Re-enable the trade button
                self.trade_button.configure(state='normal')
    
            def execute_order():
                try:
                    response = self.api_client.create_order(
                        trading_pair=order_params["trading_pair"],
                        type=order_params["type"],
                        max_amount_currency_to_trade=order_params["max_amount_currency_to_trade"],
                        price=order_params["price"],
                        end_datetime=order_params["end_datetime"],
                        min_trust_level=order_params["min_trust_level"],
                        payment_option=order_params.get("payment_option"),
                        sepa_option=order_params.get("sepa_option"),
                        seat_of_bank=order_params["seat_of_bank"]
                    )
                    if response is None:
                        raise ValueError("Received None response from API")
                    on_order_created(response)
                except Exception as e:
                    on_order_failed(e)
    
            # Execute trade through API in a separate thread
            threading.Thread(target=execute_order).start()
    
        except Exception as e:
            self.logger.error(f"Error executing trade: {str(e)}")
            messagebox.showerror("Error", f"Failed to create order: {str(e)}")
            self.trade_button.configure(state='normal')
    
    def schedule_order_check(self, order_id, order_type, max_attempts=20):
        """Schedule periodic checks for the order in the orderbook"""
        def check_order():
            nonlocal attempts
            if attempts >= max_attempts:
                
                return
                
            attempts += 1
            # Continue checking if the order hasn't been found yet
            if order_id not in self.found_orders:
                self.parent.after(1000, check_order)  # Check again in 1 second
                
        attempts = 0
        self.found_orders = getattr(self, 'found_orders', set())
        check_order()

    def show_context_menu(self, event):
        """Show context menu on right click"""
        # Get the item under cursor
        item = self.orders_tree.identify_row(event.y)
        if item:
            # Select the item
            self.orders_tree.selection_set(item)
            # Show context menu
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()
    
    def copy_order_id(self):
        """Copy the selected order ID to clipboard"""
        selected_item = self.orders_tree.selection()
        if selected_item:
            values = self.orders_tree.item(selected_item)['values']
            if values:
                order_id = values[0]
                self.parent.clipboard_clear()
                self.parent.clipboard_append(order_id)
                self.logger.debug(f"Copied order ID: {order_id}")
    
    def delete_selected_order(self):
        """Delete the selected order"""
        # Get selected item
        selected_item = self.orders_tree.selection()
        if not selected_item:
            return

        # Get order details
        values = self.orders_tree.item(selected_item)['values']
        if not values:
            return

        order_id = values[0]  # First column is Order ID
        trading_pair = values[2].lower()  # Third column is Trading Pair

        # Confirm deletion
        if not messagebox.askyesno("Bestätigen", 
                                 f"Möchten sie das Order {order_id} wirklich löschen ?"):
            return

        try:
            # Delete the order
            response = self.api_client.delete_order(trading_pair, order_id)
            
            if response and 'errors' in response and not response['errors']:
                # Remove from tree
                self.orders_tree.delete(selected_item)
                # Remove from my_orders set
                self.my_orders.discard(order_id)
                # Log success
                self.logger.info(f"Successfully deleted order {order_id}")
                messagebox.showinfo("Erfolg", "Order erfolgreich gelöscht")
            else:
                error_msg = response.get('errors', ['Unknown error'])[0] if response else 'No response'
                self.logger.error(f"Failed to delete order: {error_msg}")
                messagebox.showerror("Error", f"Failed to delete order: {error_msg}")

        except Exception as e:
            self.logger.error(f"Error deleting order: {str(e)}")
            messagebox.showerror("Error", f"Error deleting order: {str(e)}")

    def get_payment_option_value(self):
        """Convert display payment option to API value"""
        option_text = self.payment_option.get()
        if "Express & SEPA" in option_text:
            return "3"
        elif "SEPA-Only" in option_text:
            return "2"
        elif "Express-Only" in option_text:
            return "1"
        return None
       
    def get_default_end_datetime(self):
        """
        Get default end datetime (current time + 2 days)
        Returns datetime string in RFC 3339 format
        """
        # Get current time
        now = datetime.now()
    
        # Add 2 days using timedelta
        future_date = now + timedelta(days=2)
    
        # Set time to noon (12:00) of that day
        future_date = future_date.replace(hour=12, minute=0, second=0, microsecond=0)
    
        # Get timezone offset in hours
        tz_offset = time.timezone if (time.localtime().tm_isdst == 0) else time.altzone
        tz_offset = int(-tz_offset/3600)  # Convert to hours and make positive for east
    
        # Format datetime with timezone offset
        return future_date.strftime(f"%Y-%m-%dT%H:%M:00{'+' if tz_offset >= 0 else '-'}{abs(tz_offset):02d}:00")

    def update_end_datetime(self):
        """Update the end datetime automatically"""
        self.end_datetime.set(self.get_default_end_datetime())
    
        # Schedule next update for midnight
        now = datetime.now()
        tomorrow_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        delay_ms = int((tomorrow_midnight - now).total_seconds() * 1000)
    
        # Schedule next update
        self.parent.after(delay_ms, self.update_end_datetime)
    
    def calculate_total(self, *args):
        """Calculate and update total price"""
        try:
            amount = float(self.trade_amount.get())
            price = float(self.trade_price.get())
            total = amount * price
            self.total_price.set(f"{total:.2f}")
        except ValueError:
            self.total_price.set("0.00")
                     
    def validate_decimal(self, value, field_name):
        """Validate and convert a string to a Decimal"""
        try:
            return Decimal(value.replace(',', '.'))
        except InvalidOperation:
            raise InvalidOperation(f"Invalid value for {field_name}: {value}")

import openpyxl 
from openpyxl.styles import Font, Alignment
from tkinter import ttk, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from sqlite_database_manager import SQLiteDatabaseManager
from constants import Currency, CurrencyUpper
from tkinter import filedialog, messagebox
from collections import deque
import datetime
import logging

class LedgerTab:
    def __init__(self, parent, logger, api_client, db_manager: SQLiteDatabaseManager):
        self.parent = parent
        self.logger = logger
        self.api_client = api_client
        self.db_manager = db_manager
        self.setup_ui()

    def show_info_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showinfo(title, message, parent=self.root)
    
    def show_error_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showerror(title, message, parent=self.root)
    
    def show_warning_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showwarning(title, message, parent=self.root)

    def setup_ui(self):
        self.frame = ttk.Frame(self.parent)
        self.frame.grid(sticky='nsew')

        self.currency_var = tk.StringVar()
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()
        self.type_var = tk.StringVar(value="buy/sell")

        # Create selection frame
        selection_frame = ttk.Frame(self.frame)
        selection_frame.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky='ew')

        # Currency selection
        ttk.Label(selection_frame, text="Currency:").grid(row=0, column=0, padx=5, pady=5)
        all_currencies = ["alle"] + Currency.All_CURRENCIES
        self.currency_combo = ttk.Combobox(selection_frame, textvariable=self.currency_var, values=all_currencies, state="readonly")
        self.currency_combo.grid(row=0, column=1, padx=5, pady=5)

        # Date range selection
        ttk.Label(selection_frame, text="Von Datum:").grid(row=0, column=2, padx=5, pady=5)
        self.start_date_entry = ttk.Entry(selection_frame, textvariable=self.start_date_var)
        self.start_date_entry.grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(selection_frame, text="Bis Datum:").grid(row=0, column=4, padx=5, pady=5)
        self.end_date_entry = ttk.Entry(selection_frame, textvariable=self.end_date_var)
        self.end_date_entry.grid(row=0, column=5, padx=5, pady=5)

        # Type selection
        ttk.Label(selection_frame, text="Typ:").grid(row=0, column=6, padx=5, pady=5)
        self.type_combo = ttk.Combobox(selection_frame, textvariable=self.type_var, values=[
            "buy", "sell", "buy/sell", "buy/sell + kickback", "buy/sell + Kickb. + affi.", "affiliate", "kickback", "inpayment", "payout", "outgoing_fee_voluntary", "Alles"
        ], state="readonly")
        self.type_combo.grid(row=0, column=7, padx=5, pady=5)

        # Fetch button
        self.fetch_button = ttk.Button(selection_frame, text="Daten abrufen", command=self.fetch_ledger_data)
        self.fetch_button.grid(row=0, column=9, padx=5, pady=5)

        # Filter button
        self.filter_button = ttk.Button(selection_frame, text="Filtern", command=self.filter_data)
        self.filter_button.grid(row=0, column=8, padx=5, pady=5)

        # Export button
        self.export_button = ttk.Button(selection_frame, text="Exportieren nach Excel", command=self.export_to_excel)
        self.export_button.grid(row=0, column=10, padx=5, pady=5)

        # Treeview2
        self.tree = ttk.Treeview(self.frame, columns=("Currency", "Datum", "Amount", "Entry ID", "Type", "Price", "Before Fee Trade", "After Fee Trade", "Before Fee Pay", "After Fee Pay", "Balance"), show='headings')
        self.tree.heading("Currency", text="Currency")
        self.tree.heading("Datum", text="Datum")
        self.tree.heading("Amount", text="Amount")
        self.tree.heading("Entry ID", text="Entry ID")
        self.tree.heading("Type", text="Type")
        self.tree.heading("Price", text="Price")
        self.tree.heading("Before Fee Trade", text="Before Fee Trade")
        self.tree.heading("After Fee Trade", text="After Fee Trade")
        self.tree.heading("Before Fee Pay", text="Before Fee Pay")
        self.tree.heading("After Fee Pay", text="After Fee Pay")
        self.tree.heading("Balance", text="Balance")
        self.tree.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky='nsew')

        # Set column widths
        self.tree.column("Currency", width=10)
        self.tree.column("Datum", width=45)
        self.tree.column("Amount", width=45)
        self.tree.column("Entry ID", width=45)
        self.tree.column("Type", width=45)
        self.tree.column("Price", width=45)
        self.tree.column("Before Fee Trade", width=45)
        self.tree.column("After Fee Trade", width=45)
        self.tree.column("Before Fee Pay", width=45)
        self.tree.column("After Fee Pay", width=45)
        self.tree.column("Balance", width=45)

        # Add total profit/loss label
        self.total_profit_loss_label = ttk.Label(self.frame, text="Total Gewinn/Verlust: 0.00 €")
        self.total_profit_loss_label.grid(row=2, column=0, columnspan=3, padx=10, pady=10, sticky='w')

    def fetch_ledger_data(self):
        currencies = Currency.All_CURRENCIES
        try:
            all_data = self.api_client.get_all_account_ledgers(currencies)
            
            for currency, data in all_data.items():
                self.db_manager.save_ledger_data(currency, {'account_ledger': data})
            messagebox.showinfo("Erfolg", "Daten erfolgreich abgerufen und gespeichert")
        except Exception as e:
            self.logger.error(f"Error fetching ledger data: {str(e)}")
            messagebox.showerror("Fehler", "Fehler beim Abrufen der Daten. Siehe Protokoll für Details.")
       
    def calculate_fifo(self, rows):
        total_buy_before_fee_pay = 0.0
        total_sell_after_fee_pay = 0.0
        total_fees_paid = 0.0
        total_fifo_profit_loss = 0.0
        fifo_queue = deque()  # FIFO queue for tracking buys
    
        # Sort rows by timestamp in ascending order to start with the oldest entry
        rows.sort(key=lambda x: datetime.strptime(x[3], "%Y-%m-%dT%H:%M:%S%z"))
    
        for row in rows:
            amount = abs(float(row[2]))  # Use absolute value of amount
            price = float(row[5])
            before_fee_trade = float(row[6])
            after_fee_trade = float(row[7])
            before_fee_pay = float(row[8])
            after_fee_pay = float(row[9])
            entry_type = row[4].lower()
            timestamp = datetime.strptime(row[3], "%Y-%m-%dT%H:%M:%S%z")
            fifo_profit_loss = 0.0
    
            if entry_type in ['buy', 'kickback', 'affiliate','initial_fork']:
                fee_paid = (before_fee_trade - after_fee_trade) * price
                total_buy_before_fee_pay += before_fee_pay
                total_cost = amount * price  # Gesamtbetrag vor Gebühren
                fifo_queue.append((amount, total_cost, timestamp))  # Add to FIFO queue with timestamp
                logging.debug(f"BUY: amount={amount}, total_cost={total_cost}, timestamp={timestamp}")
            elif entry_type == 'sell':
                fee_paid = before_fee_pay - after_fee_pay
                total_sell_after_fee_pay += after_fee_pay
                # Sort the FIFO queue by timestamp
                fifo_queue = deque(sorted(fifo_queue, key=lambda x: x[2]))
                # Calculate FIFO profit/loss
                remaining_amount = amount
                while remaining_amount > 0 and fifo_queue:
                    buy_amount, buy_total_cost, buy_timestamp = fifo_queue.popleft()
                    holding_period = (timestamp - buy_timestamp).days
                    logging.debug(f"SELL: remaining_amount={remaining_amount}, buy_amount={buy_amount}, buy_total_cost={buy_total_cost}, buy_timestamp={buy_timestamp}, holding_period={holding_period}")
                    if holding_period > 365:
                        # Coins held for more than a year are not taxable
                        fifo_profit_loss += 0
                    else:
                        if buy_amount <= remaining_amount:
                            fifo_profit_loss += (price - (buy_total_cost / buy_amount)) * buy_amount
                            remaining_amount -= buy_amount
                        else:
                            fifo_profit_loss += (price - (buy_total_cost / buy_amount)) * remaining_amount
                            fifo_queue.appendleft((buy_amount - remaining_amount, (buy_total_cost / buy_amount) * (buy_amount - remaining_amount), buy_timestamp))
                            remaining_amount = 0
            else:
                fee_paid = 0.0
            total_fees_paid += fee_paid
            total_fifo_profit_loss += fifo_profit_loss
    
        total_profit_loss = total_fifo_profit_loss - total_fees_paid
        return total_buy_before_fee_pay, total_sell_after_fee_pay, total_fees_paid, total_fifo_profit_loss, total_profit_loss
    
    def filter_data(self):
        try:
            currency = self.currency_var.get()
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()
            entry_type = self.type_var.get()
    
            query = "SELECT currency, entry_id, amount, timestamp, type, price, before_fee_trade, after_fee_trade, before_fee_pay, after_fee_pay, balance FROM ledger WHERE 1=1"
            params = []
    
            if currency and currency != "alle":
                query += " AND currency = ?"
                params.append(currency)
    
            if start_date:
                if len(start_date) == 4:  # Year only
                    start_date = f"{start_date}-01-01"
                else:
                    start_date = datetime.strptime(start_date, "%d.%m.%Y").strftime("%Y-%m-%d")
                query += " AND timestamp >= ?"
                params.append(start_date)
    
            if end_date:
                if len(end_date) == 4:  # Year only
                    end_date = f"{end_date}-12-31"
                else:
                    end_date = datetime.strptime(end_date, "%d.%m.%Y").strftime("%Y-%m-%d")
                query += " AND timestamp <= ?"
                params.append(end_date)
    
            if entry_type == "buy":
                query += " AND type = 'buy'"
            elif entry_type == "sell":
                query += " AND type = 'sell'"
            elif entry_type == "buy/sell":
                query += " AND (type = 'buy' OR type = 'sell')"
            elif entry_type == "buy/sell + kickback":
                query += " AND type IN ('buy', 'sell', 'kickback')"
            elif entry_type == "buy/sell + Kickb. + affi.":
                query += " AND type IN ('buy', 'sell', 'kickback', 'affiliate')"
            elif entry_type == "affiliate":
                query += " AND type = 'affiliate'"
            elif entry_type == "kickback":
                query += " AND type = 'kickback'"
            elif entry_type == "inpayment":
                query += " AND type = 'inpayment'"
            elif entry_type == "payout":
                query += " AND type = 'payout'"
            elif entry_type == "outgoing_fee_voluntary":
                query += " AND type = 'outgoing_fee_voluntary'"
            elif entry_type == "Alles":
                query += " AND type IN ('buy', 'sell', 'affiliate', 'kickback', 'inpayment', 'payout', 'outgoing_fee_voluntary', 'initial', 'initial_fork', 'manual_increase')"

            conn = self.db_manager.get_sqlite_connection()
            cursor = conn.cursor()
            cursor.execute(query, params)
            rows = cursor.fetchall()
            cursor.close()
            conn.close()
    
            # Sort rows by timestamp in descending order
            rows.sort(key=lambda x: datetime.strptime(x[3], "%Y-%m-%dT%H:%M:%S%z"), reverse=True)
    
            for row in self.tree.get_children():
                self.tree.delete(row)
    
            for row in rows:
                formatted_row = (
                    row[0],  # currency
                    datetime.strptime(row[3], "%Y-%m-%dT%H:%M:%S%z").strftime("%d.%m.%Y"),  # timestamp
                    row[2],  # amount
                    row[1],  # entry_id
                    row[4],  # type
                    row[5],  # price
                    row[6],  # before_fee_trade
                    row[7],  # after_fee_trade
                    row[8],  # before_fee_pay
                    row[9],  # after_fee_pay
                    row[10]  # balance
                )
                self.tree.insert('', 'end', values=formatted_row)
    
            # FIFO-Berechnung ausführen und relevante Werte anzeigen
            total_buy_before_fee_pay, total_sell_after_fee_pay, total_fees_paid, total_fifo_profit_loss, total_profit_loss = self.calculate_fifo(rows)
            logging.debug(f"Total Buy Before Fee Pay: {total_buy_before_fee_pay}")
            logging.debug(f"Total Sell After Fee Pay: {total_sell_after_fee_pay}")
            logging.debug(f"Total Fees Paid: {total_fees_paid}")
            logging.debug(f"Total FIFO Profit/Loss: {total_fifo_profit_loss}")
            logging.debug(f"Total Profit/Loss: {total_profit_loss}")
    
            # Format the total profit/loss with two decimal places
            formatted_total_profit_loss = f"{total_profit_loss:.2f} €"
    
            # Update total profit/loss label
            self.total_profit_loss_label.config(text=f"Total Gewinn/Verlust: {formatted_total_profit_loss}")
    
        except Exception as e:
            self.logger.error(f"Error filtering data: {str(e)}")
            messagebox.showerror("Fehler", "Fehler beim Filtern der Daten. Siehe Protokoll für Details.")
    
    def export_to_excel(self):
        logging.info("Starting export to Excel")
        try:
            # Ask the user for the file path to save the Excel file
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                logging.debug("Export cancelled by user")
                return
    
            # Get the selected currency for naming the sheet
            selected_currency = self.currency_var.get()
            if selected_currency == "alle":
                selected_currency = "All_Currencies"
                currency_symbol = "EUR"
            else:
                currency_symbol = CurrencyUpper.All_CURRENCIES[Currency.All_CURRENCIES.index(selected_currency.lower())]
    
            logging.info(f"Selected currency: {selected_currency}, Currency symbol: {currency_symbol}")
    
            # Create a new Excel workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"{selected_currency}"
    
            # Write header
            headers = [
                "Währung", "Datum", "Typ", "Order-ID", "Menge", "Preis", 
                f"{currency_symbol} v. Gebühr", f"{currency_symbol} n. Gebühr", 
                "€ v. Gebühr", "€ n. Gebühr", "Saldo", "Gebühren €", "FIFO GuV"
            ]
            sheet.append(headers)
            logging.debug("Headers written to Excel sheet")
    
            # Set column widths
            column_widths = [10, 10, 10, 10, 15, 10, 18, 18, 12, 12, 19, 13, 13]
            for col_num, width in enumerate(column_widths, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
            # Format header
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center')
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = header_alignment
    
            total_buy_before_fee_pay = 0.0
            total_sell_after_fee_pay = 0.0
            total_fees_paid = 0.0
            fifo_queue = deque()  # FIFO queue for tracking buys
    
            # Get all items from the treeview in reverse order
            tree_items = self.tree.get_children()[::-1]
            logging.debug(f"Number of items to export: {len(tree_items)}")
    
            # Write data to the sheet
            for row in tree_items:
                values = self.tree.item(row, 'values')
                formatted_values = []
                fifo_profit_loss = 0.0
                for i, value in enumerate(values):
                    try:
                        # Attempt to convert to float if the value is numeric
                        numeric_value = float(value)
                        formatted_values.append(numeric_value)
                    except ValueError:
                        # If conversion fails, keep the value as is (likely a date or non-numeric string)
                        formatted_values.append(value)
    
                # Calculate the paid fees in €
                try:
                    amount = abs(float(values[2]))  # Use absolute value of amount
                    price = float(values[5])
                    before_fee_trade = float(values[6])
                    after_fee_trade = float(values[7])
                    before_fee_pay = float(values[8])
                    after_fee_pay = float(values[9])
                    entry_type = values[4].lower()
                    timestamp = datetime.strptime(values[1], "%d.%m.%Y")
                    if entry_type in ['buy', 'kickback', 'affiliate', 'initial_fork']:
                        fee_paid = (before_fee_trade - after_fee_trade) * price
                        total_buy_before_fee_pay += before_fee_pay
                        total_cost = amount * price  # Gesamtbetrag vor Gebühren
                        fifo_queue.append((amount, total_cost, timestamp))  # Add to FIFO queue with timestamp
                        logging.debug(f"BUY: amount={amount}, total_cost={total_cost}, timestamp={timestamp}")
                    elif entry_type == 'sell':
                        fee_paid = before_fee_pay - after_fee_pay
                        total_sell_after_fee_pay += after_fee_pay
                        # Sort the FIFO queue by timestamp
                        fifo_queue = deque(sorted(fifo_queue, key=lambda x: x[2]))
                        # Calculate FIFO profit/loss
                        remaining_amount = amount
                        while remaining_amount > 0 and fifo_queue:
                            buy_amount, buy_total_cost, buy_timestamp = fifo_queue.popleft()
                            holding_period = (timestamp - buy_timestamp).days
                            logging.debug(f"SELL: remaining_amount={remaining_amount}, buy_amount={buy_amount}, buy_total_cost={buy_total_cost}, buy_timestamp={buy_timestamp}, holding_period={holding_period}")
                            if holding_period > 365:
                                # Coins held for more than a year are not taxable
                                fifo_profit_loss += 0
                            else:
                                if buy_amount <= remaining_amount:
                                    fifo_profit_loss += (price - (buy_total_cost / buy_amount)) * buy_amount
                                    remaining_amount -= buy_amount
                                else:
                                    fifo_profit_loss += (price - (buy_total_cost / buy_amount)) * remaining_amount
                                    fifo_queue.appendleft((buy_amount - remaining_amount, (buy_total_cost / buy_amount) * (buy_amount - remaining_amount), buy_timestamp))
                                    remaining_amount = 0
                    else:
                        fee_paid = 0.0
                    formatted_values.append(fee_paid)
                    total_fees_paid += fee_paid
                except (ValueError, IndexError):
                    formatted_values.append(0.0)
    
                # Add FIFO profit/loss to formatted values
                formatted_values.append(fifo_profit_loss)
    
                # Reorder columns: E (Amount) to C, C (Type) to E
                reordered_values = [
                    formatted_values[0],  # Currency
                    formatted_values[1],  # Datum
                    formatted_values[4],  # Typ (was Menge)
                    formatted_values[3],  # Order-ID
                    formatted_values[2],  # Menge (was Typ)
                    formatted_values[5],  # Preis
                    formatted_values[6],  # Vor Gebühr Handel
                    formatted_values[7],  # Nach Gebühr Handel
                    formatted_values[8],  # Vor Gebühr Zahlung
                    formatted_values[9],  # Nach Gebühr Zahlung
                    formatted_values[10], # Saldo
                    formatted_values[11], # Gebühren in €
                    formatted_values[12]  # FIFO Gewinn/Verlust
                ]
    
                sheet.append(reordered_values)
    
            logging.debug("Data rows written to Excel sheet")
    
            # Format data rows
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    if cell.column in [1, 2, 3, 4]:  # Columns A, B, C, D
                        cell.alignment = Alignment(horizontal='center')
                    else:  # Columns E, F, G, H, I, J, K, L, M
                        cell.alignment = Alignment(horizontal='right')
    
            # Add summary rows
            last_row = sheet.max_row + 2
            summary_font = Font(bold=True, size=13)
            sheet.merge_cells(f"D{last_row}:E{last_row}")
            sheet[f"D{last_row}"] = "Alle Käufe"
            sheet[f"D{last_row}"].alignment = Alignment(horizontal='center')
            sheet[f"D{last_row}"].font = summary_font
            sheet.merge_cells(f"D{last_row + 1}:E{last_row + 1}")
            sheet[f"D{last_row + 1}"] = f"{total_buy_before_fee_pay:.2f} €"
            sheet[f"D{last_row + 1}"].alignment = Alignment(horizontal='center')
            sheet[f"D{last_row + 1}"].font = summary_font
    
            sheet.merge_cells(f"F{last_row}:G{last_row}")
            sheet[f"F{last_row}"] = "Alle Verkäufe"
            sheet[f"F{last_row}"].alignment = Alignment(horizontal='center')
            sheet[f"F{last_row}"].font = summary_font
            sheet.merge_cells(f"F{last_row + 1}:G{last_row + 1}")
            sheet[f"F{last_row + 1}"] = f"{total_sell_after_fee_pay:.2f} €"
            sheet[f"F{last_row + 1}"].alignment = Alignment(horizontal='center')
            sheet[f"F{last_row + 1}"].font = summary_font
    
            sheet[f"H{last_row}"] = "€ in Umlauf"
            sheet[f"H{last_row}"].alignment = Alignment(horizontal='center')
            sheet[f"H{last_row}"].font = summary_font
            sheet[f"H{last_row + 1}"] = f"{total_sell_after_fee_pay - total_buy_before_fee_pay - total_fees_paid:.2f} €"
            sheet[f"H{last_row + 1}"].alignment = Alignment(horizontal='center')
            sheet[f"H{last_row + 1}"].font = summary_font
    
            # Ermitteln des letzten nicht-leeren Werts in der Spalte K (Saldo)
            last_balance_value = None
            for row in range(sheet.max_row, 0, -1):  # Von der letzten Zeile nach oben iterieren
                cell_value = sheet[f"K{row}"].value
                if cell_value is not None:  # Prüfen, ob die Zelle nicht leer ist
                    last_balance_value = cell_value
                    break
            
            # Aktualisieren des Bereichs "Bestand {currency_symbol}"
            sheet.merge_cells(f"I{last_row}:J{last_row}")
            sheet[f"I{last_row}"] = f"Bestand {currency_symbol}"
            sheet[f"I{last_row}"].alignment = Alignment(horizontal='center')
            sheet[f"I{last_row}"].font = summary_font
            
            sheet.merge_cells(f"I{last_row + 1}:J{last_row + 1}")
            sheet[f"I{last_row + 1}"] = last_balance_value  # Verwenden des gefundenen Werts
            sheet[f"I{last_row + 1}"].alignment = Alignment(horizontal='center')
            sheet[f"I{last_row + 1}"].font = summary_font
    
            # Calculate total FIFO profit/loss
            total_fifo_profit_loss = 0.0
            for row in sheet.iter_rows(min_row=2, max_row=last_row - 2, min_col=13, max_col=13, values_only=True):
                if row and len(row) == 1 and row[0] is not None:
                    total_fifo_profit_loss += float(row[0])
    
            # Add total FIFO profit/loss to summary
            sheet[f"K{last_row}"] = "Total FIFO GuV"
            sheet[f"K{last_row}"].alignment = Alignment(horizontal='center')
            sheet[f"K{last_row}"].font = summary_font
            sheet[f"K{last_row + 1}"] = f"{total_fifo_profit_loss:.2f} €"
            sheet[f"K{last_row + 1}"].alignment = Alignment(horizontal='center')
            sheet[f"K{last_row + 1}"].font = Font(bold=True, size=13)
    
            # Calculate total profit/loss
            total_profit_loss = total_fifo_profit_loss - total_fees_paid
    
            # Add total profit/loss to summary with one row gap
            sheet.merge_cells(f"L{last_row + 3}:M{last_row + 3}")
            sheet[f"L{last_row + 3}"] = "Total Gewinn/Verlust"
            sheet[f"L{last_row + 3}"].alignment = Alignment(horizontal='center')
            sheet[f"L{last_row + 3}"].font = summary_font
            sheet.merge_cells(f"L{last_row + 4}:M{last_row + 4}")
            sheet[f"L{last_row + 4}"] = f"{total_profit_loss:.2f} €"
            sheet[f"L{last_row + 4}"].alignment = Alignment(horizontal='center')
            sheet[f"L{last_row + 4}"].font = Font(bold=True, size=13, underline='double')
    
            # Add total fees to summary
            sheet.merge_cells(f"L{last_row}:M{last_row}")
            sheet[f"L{last_row}"] = "Alle Gebühren"
            sheet[f"L{last_row}"].alignment = Alignment(horizontal='left')
            sheet[f"L{last_row}"].font = summary_font
            sheet.merge_cells(f"L{last_row + 1}:M{last_row + 1}")
            sheet[f"L{last_row + 1}"] = f"{total_fees_paid:.2f} €"
            sheet[f"L{last_row + 1}"].alignment = Alignment(horizontal='left')
            sheet[f"L{last_row + 1}"].font = summary_font
    
            # Add explanation text
            explanation_row = last_row + 6
            sheet.merge_cells(f"A{explanation_row}:I{explanation_row + 21}")
            explanation_text = (
                "Erläuterungen:\n\n"
                "Die Gebühren für Käufe werden wie folgt berechnet: "
                f"Kaufgebühr = ({currency_symbol} v. Gebühr - {currency_symbol} n. Gebühr) x Preis\n"
                "Die Gebühren für Verkäufe werden wie folgt berechnet: "
                "Verkaufsgebühr = € v. Gebühr - € n. Gebühr\n\n"
                f"Kickbacks sind Rückvergütungen in {currency_symbol} der gezahlten Marktplatzgebühren, diese werden in der " 
                "FIFO-Berechnung mit einem Einkaufswert von 0€ einbezogen.\n"
                f"Affiliate sind Provisionen in {currency_symbol} für das Werben von Kunden, diese werden ebenfalls mit 0€ berechnet.\n\n"
                "Die FIFO (First In, First Out) Berechnungsmethode bedeutet, dass die zuerst gekauften Einheiten einer Währung "
                "auch die ersten sind, die verkauft werden. Dies wird verwendet, um den Gewinn oder Verlust aus Verkäufen zu berechnen, "
                "indem die Kosten der zuerst gekauften Einheiten von den Verkaufserlösen abgezogen werden. Diese Berechnung ist für "
                f"Privatpersonen, deshalb wird der Gewinn aus einem Verkauf von {currency_symbol}, die länger als ein Jahr gehalten wurden, "
                "automatisch nicht als Gewinn berechnet. "
                "Der FIFO-Gewinn/Verlust wird automatisch berechnet und bezieht alle Faktoren mit ein.\n"
                "Beispiel:\n"
                "1. Kauf von 10 Einheiten zu 100 € am 01.01.2025,\n"
                "2. Kauf von 5 Einheiten zu 120 € am 01.02.2025,\n"
                "3. Verkauf von 8 Einheiten zu 150 € am 01.03.2025\n"
                "Verkaufserlös: 8 x 150 € = 1200 €\n"
                "FIFO-Berechnung = 10 zu 100€: 8 Einheiten werden verkauft, verbleiben 2 Einheiten\n"
                "FIFO-Kosten = 8 x 100€ = 800€\n"
                "FIFO-Gewinn/Verlust = 1200€ - 800€ = 400€\n"
            )
            sheet[f"A{explanation_row}"] = explanation_text
            sheet[f"A{explanation_row}"].alignment = Alignment(wrap_text=True, vertical='top')
    
            # Format "Erläuterungen:" as bold
            sheet[f"A{explanation_row}"].font = Font(bold=True)
    
            # Format "Käufe" and "Verkäufe" as bold
            sheet[f"A{explanation_row + 2}"].font = Font(bold=True)
            sheet[f"A{explanation_row + 6}"].font = Font(bold=True)
    
            # Set page setup options
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    
            # Save the workbook
            workbook.save(file_path)
            logging.debug(f"Workbook saved to {file_path}")
    
            messagebox.showinfo("Erfolg", "Daten erfolgreich exportiert")
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {str(e)}")
            messagebox.showerror("Fehler", f"Fehler beim Exportieren nach Excel: {str(e)}")

from tkinter import ttk, StringVar, IntVar, BooleanVar, messagebox
from ttkbootstrap.constants import *
from datetime import datetime, timedelta, timezone
import time
from constants import TradingPairs  
import threading

class TradeBotTab:
    def __init__(self, parent, logger, api_client, db_manager):
        self.parent = parent
        self.logger = logging.getLogger("TradeBotLogger")  # Neuer Logger für den Trade-Bot
        self.logger.setLevel(logging.INFO)  # Setze das Log-Level
        self.api_client = api_client
        self.db_manager = db_manager
    
        # Variablen für die Einstellungen des Bots
        self.selected_pair = StringVar(value=TradingPairs.DISPLAY_NAMES[TradingPairs.BTCEUR])
        self.trade_amount = StringVar(value="0.00")
        self.trade_price = StringVar(value="00.00")
        self.total_price = StringVar(value="0.00")
        self.trade_type = StringVar(value="buy")
        self.max_price = StringVar(value=TradingConstants.DEFAULT_MAX_PRICE)
        self.min_price = StringVar(value=TradingConstants.DEFAULT_MIN_PRICE)
        self.end_datetime = StringVar(value=self.get_default_end_datetime())
        self.min_trust_level = StringVar(value=TradingConstants.TrustLevel.BRONZE.value)
        self.payment_option = IntVar(value=TradingConstants.ORDERBOOK_PAYMENT_OPTION_SEPA_ONLY)
        self.sepa_option = IntVar(value=0)
        self.seat_of_bank = StringVar(value=TradingConstants.BankCountry.GERMANY.value)
        self.auto_trade = BooleanVar(value=False)
    
        # Mappings für Handelspaare
        self.pair_mapping = {display: api_name for api_name, display in TradingPairs.DISPLAY_NAMES.items()}

        # Mappings für Optionen
        self.payment_options_map = {
            "Express & SEPA": 3,
            "SEPA-Only": 2,
            "Express-Only": 1
        }
        self.sepa_options_map = {
            "SEPA Instant": 1,
            "SEPA Normal": 0
        }
    
        # Automatische Berechnung des Gesamtpreises
        self.trade_amount.trace('w', self.calculate_total)
        self.trade_price.trace('w', self.calculate_total)
    
        self.setup_ui()

    def setup_ui(self):
        # Create main frame
        self.main_frame = ttk.Frame(self.parent)
        self.main_frame.pack(fill='both', expand=True)
    
        # Create controls frame
        self.controls_frame = ttk.LabelFrame(self.main_frame, text="Trade Einstellungen")
        self.controls_frame.pack(fill='x', padx=5, pady=(5, 0))
    
        # Create trading pair selection frame
        pair_frame = ttk.Frame(self.controls_frame)
        pair_frame.pack(fill='x', padx=1, pady=5)
    
        # Add trading pair label and dropdown
        ttk.Label(pair_frame, text="Handelspaar:").pack(side='left', padx=1)
        self.pair_dropdown = ttk.Combobox(
            pair_frame,
            textvariable=self.selected_pair,
            values=TradingPairs.get_all_display_names(),
            state="readonly",
            width=9
        )
        self.pair_dropdown.pack(side='left', padx=1)
    
        # Add trade amount entry
        ttk.Label(pair_frame, text="Menge:").pack(side='left', padx=1)
        self.amount_entry = ttk.Entry(pair_frame, textvariable=self.trade_amount, width=7)
        self.amount_entry.pack(side='left', padx=1)
    
        # Add min price entry
        ttk.Label(pair_frame, text="Min Preis:").pack(side='left', padx=1)
        self.min_price_entry = ttk.Entry(pair_frame, textvariable=self.min_price, width=7)
        self.min_price_entry.pack(side='left', padx=1)

        # Add max price entry
        ttk.Label(pair_frame, text="Max Preis:").pack(side='left', padx=1)
        self.max_price_entry = ttk.Entry(pair_frame, textvariable=self.max_price, width=8)
        self.max_price_entry.pack(side='left', padx=1)

        # Add total price display
        ttk.Label(pair_frame, text="Total (€):").pack(side='left', padx=1)
        self.total_label = ttk.Label(pair_frame, textvariable=self.total_price)
        self.total_label.pack(side='left', padx=1)
    
        # Add trade type selection with binding
        ttk.Label(pair_frame, text="Type:").pack(side='left', padx=1)
        self.type_dropdown = ttk.Combobox(
            pair_frame,
            textvariable=self.trade_type,
            values=["buy", "sell"],
            state="readonly",
            width=4
        )
        self.type_dropdown.pack(side='left', padx=1)
        self.type_dropdown.bind('<<ComboboxSelected>>', self.on_trade_type_changed)
    
        # Payment option selection frame
        self.payment_frame = ttk.Frame(pair_frame)
        self.payment_frame.pack(side='left')
    
        ttk.Label(self.payment_frame, text="Bezahl Option:").pack(side='left', padx=1)
        self.payment_option_combo = ttk.Combobox(
            self.payment_frame,
            values=list(self.payment_options_map.keys()),
            state="readonly",
            width=14
        )
        self.payment_option_combo.pack(side='left', padx=1)
        self.payment_option_combo.bind('<<ComboboxSelected>>', self.on_payment_option_changed)
    
        # SEPA option selection frame
        self.sepa_frame = ttk.Frame(pair_frame)
        self.sepa_frame.pack(side='left')
    
        ttk.Label(self.sepa_frame, text="SEPA Option:").pack(side='left', padx=1)
        self.sepa_option_combo = ttk.Combobox(
            self.sepa_frame,
            values=list(self.sepa_options_map.keys()),
            state="readonly",
            width=15
        )
        self.sepa_option_combo.pack(side='left', padx=1)
        self.sepa_option_combo.bind('<<ComboboxSelected>>', self.on_sepa_option_changed)
    
        # Add trust level selection
        ttk.Label(pair_frame, text="Trust Level:").pack(side='left', padx=1)
        self.trust_level_dropdown = ttk.Combobox(
            pair_frame,
            textvariable=self.min_trust_level,
            values=[level.value for level in TradingConstants.TrustLevel],
            state="readonly",
            width=7
        )
        self.trust_level_dropdown.pack(side='left', padx=1)

        # Bank country selection
        ttk.Label(pair_frame, text="Sitz der Bank:").pack(side='left', padx=1)
        self.bank_country_combo = ttk.Combobox(
            pair_frame,
            textvariable=self.seat_of_bank,
            values=[f"{code} {name}" for code, name in TradingConstants.BANK_COUNTRY_NAMES.items()],
            state="readonly",
            width=5
        )
        self.bank_country_combo.pack(side='left', padx=1)
    
        # Add end datetime entry with auto-update
        ttk.Label(pair_frame, text="Aktiv bis:").pack(side='left', padx=1)
        self.end_datetime_entry = ttk.Entry(pair_frame, textvariable=self.end_datetime, width=25)
        self.end_datetime_entry.pack(side='left', padx=1)
    
        # Initialize the UI state
        self.trade_type.set("buy")  # Set default to buy
        self.on_trade_type_changed(None)  # Update UI state
    
        # Start automatic updates
        self.update_end_datetime()

        # Start-/Stop-Button
        self.start_button = ttk.Button(self.controls_frame, text="Bot Starten", command=self.start_bot)
        self.start_button.pack(side='left', padx=5, pady=10)

        self.stop_button = ttk.Button(self.controls_frame, text="Bot Stoppen", command=self.stop_bot, state="disabled")
        self.stop_button.pack(side='left', padx=5, pady=10)

        # Logging-Konsole hinzufügen
        console_frame = ttk.LabelFrame(self.main_frame, text="TradeBot Logs", padding="5")
        console_frame.pack(fill='both', expand=True, padx=5, pady=5)
    
        # Text-Widget für die Konsole
        self.console = tk.Text(console_frame, height=10, wrap='word', state='disabled')
        self.console.pack(fill='both', expand=True, padx=5, pady=5)
    
        # Add time settings frame
        time_settings_frame = ttk.LabelFrame(self.main_frame, text="Zeiteinstellungen")
        time_settings_frame.pack(fill='x', padx=5, pady=5)

        # Add check interval setting
        ttk.Label(time_settings_frame, text="Überprüfungsintervall (Sekunden):").pack(side='left', padx=5)
        self.check_interval = IntVar(value=30)  # Standardwert: 15 Sekunden
        self.check_interval_entry = ttk.Entry(time_settings_frame, textvariable=self.check_interval, width=5)
        self.check_interval_entry.pack(side='left', padx=5)

        # Add order interval setting
        ttk.Label(time_settings_frame, text="Orderintervall (Sekunden):").pack(side='left', padx=5)
        self.order_interval = IntVar(value=15)  # Standardwert: 5 Sekunden
        self.order_interval_entry = ttk.Entry(time_settings_frame, textvariable=self.order_interval, width=5)
        self.order_interval_entry.pack(side='left', padx=5)
    
        # Logging-Handler hinzufügen
        self.add_console_handler()

    def add_console_handler(self):
        """Fügt einen benutzerdefinierten Logging-Handler für die TradeBot-Konsole hinzu."""
        class ConsoleHandler(logging.Handler):
            def __init__(self, console):
                super().__init__()
                self.console = console
    
            def emit(self, record):
                msg = self.format(record) + '\n'
                self.console.configure(state='normal')  # Aktivieren des Schreibens
    
                # Schlüsselwörter, die fett dargestellt werden sollen
                keywords = ["Höchster Ankaufs-Preis",
                            "Niedrigster Verkaufs-Preis", 
                            "Order erfolgreich erstellt", 
                            "Aktueller Preis",
                            "Nächstniedriger Preis im Orderbuch",
                            "Nächsthöherer Preis im Orderbuch",
                            "Lösche bestehendes Order",
                            "Order erfolgreich gelöscht", 
                            "neuer Preis"]
    
                # Prüfen, ob die Nachricht eines der Schlüsselwörter enthält
                for keyword in keywords:
                    if keyword in msg:
                        # Fett formatieren
                        self.console.tag_configure("bold", font=("Helvetica", 10, "bold"))
    
                        # Teile die Nachricht in normale und fett formatierte Teile
                        parts = msg.split(": ")
                        if len(parts) == 2:
                            # Erster Teil (normaler Text)
                            self.console.insert('end', parts[0] + ": ")
    
                            # Zweiter Teil (fett formatierter Wert)
                            self.console.insert('end', parts[1], "bold")
                        else:
                            # Falls die Nachricht nicht wie erwartet formatiert ist, normal einfügen
                            self.console.insert('end', msg)
                        break
                else:
                    # Normale Nachricht
                    self.console.insert('end', msg)
    
                self.console.configure(state='disabled')  # Deaktivieren des Schreibens
                self.console.see('end')  # Scrollen zum Ende
    
        # Erstellen und Konfigurieren des Handlers
        console_handler = ConsoleHandler(self.console)
        console_handler.setLevel(logging.DEBUG)  # Nur INFO und höher loggen
        console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    
        # Hinzufügen des Handlers zum TradeBotLogger
        self.logger.addHandler(console_handler)

    def on_trade_type_changed(self, _):
        """Handle trade type changes"""
        trade_type = self.trade_type.get()
        
        if trade_type == "sell":
            # Show payment options for sell orders
            self.payment_frame.pack(side='left')
            self.payment_option_combo.configure(state="readonly")
            
            # Set default to SEPA-Only
            self.payment_option_combo.set("SEPA-Only")
            self.payment_option.set(2)  # Set the IntVar to SEPA-Only (2)
            
            # Set default SEPA option to Instant
            self.sepa_frame.pack(side='left')
            self.sepa_option_combo.set("SEPA Instant")
            self.sepa_option.set(1)  # Set to SEPA Instant (1)
            
            # Update SEPA options state
            self.on_payment_option_changed(None)
        else:
            # Hide/disable payment options for buy orders
            self.payment_frame.pack_forget()
            self.sepa_frame.pack_forget()
            self.payment_option.set(0)
            self.sepa_option.set(0)

    def on_payment_option_changed(self, _):
        """Handle payment option changes"""
        if self.trade_type.get() == "sell":
            selected_text = self.payment_option_combo.get()
            payment_value = self.payment_options_map.get(selected_text, 0)
            self.payment_option.set(payment_value)  # Update the IntVar
            
            # Enable SEPA options only for SEPA-Only (2) or Express & SEPA (3)
            if payment_value in [2, 3]:
                self.sepa_frame.pack(side='left')
                self.sepa_option_combo.configure(state="readonly")
                # Set default to SEPA Instant if not already set
                if not self.sepa_option_combo.get():
                    self.sepa_option_combo.set("SEPA Instant")
                    self.sepa_option.set(1)
            else:
                self.sepa_frame.pack_forget()
                self.sepa_option.set(0)

    def on_sepa_option_changed(self, _):
        """Handle SEPA option changes"""
        selected_text = self.sepa_option_combo.get()
        sepa_value = self.sepa_options_map.get(selected_text, 0)
        self.sepa_option.set(sepa_value)  # Update the IntVar with numeric value

    def update_end_datetime(self):
        """Update the end datetime automatically"""
        self.end_datetime.set(self.get_default_end_datetime())
    
        # Schedule next update for midnight
        now = datetime.now()
        tomorrow_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        delay_ms = int((tomorrow_midnight - now).total_seconds() * 1000)
    
        # Schedule next update
        self.parent.after(delay_ms, self.update_end_datetime)

    def calculate_total(self, *args):
        """Berechnet den Gesamtpreis."""
        try:
            amount = float(self.trade_amount.get())
            price = float(self.min_price.get())
            total = amount * price
            self.total_price.set(f"{total:.2f}")
        except ValueError:
            self.total_price.set("0.00")

    def get_default_end_datetime(self):
        """
        Gibt das Standard-Enddatum zurück (2 Tage in der Zukunft) im RFC 3339-Format mit Zeitzone.
        """
        now = datetime.now()
    
        future_date = now + timedelta(days=2)
    
        future_date = future_date.replace(hour=12, minute=0, second=0, microsecond=0)
    
        tz_offset = time.localtime().tm_gmtoff  # Offset in Sekunden
        tz = timezone(timedelta(seconds=tz_offset))  # Zeitzone erstellen
        future_date_with_tz = future_date.replace(tzinfo=tz)
    
        # Formatieren im RFC 3339-Format
        return future_date_with_tz.isoformat()
  
    def start_bot(self):
        """Startet den Trade-Bot mit Sicherheitsabfrage."""
        # Sammle die aktuellen Einstellungen
        settings_summary = (
            f"Handelspaar: {self.selected_pair.get()}\n"
            f"Typ: {self.trade_type.get()}\n"
            f"Menge: {self.trade_amount.get()}\n"
            f"Min Preis: {self.min_price.get()}\n"
            f"Max Preis: {self.max_price.get()}\n"            
            f"Enddatum: {self.end_datetime.get()}\n"
            f"Trust Level: {self.min_trust_level.get()}\n"
            f"Bezahloption: {self.payment_option_combo.get()}\n"
            f"SEPA Option: {self.sepa_option_combo.get()}\n"
            f"Sitz der Bank: {self.seat_of_bank.get()}"
        )
    
        # Zeige die Sicherheitsabfrage
        confirm = messagebox.askyesno(
            "Eingabewerte überprüfen",
            f"Bitte überprüfen Sie die folgenden Einstellungen:\n\n{settings_summary}\n\nSind diese korrekt?"
        )
    
        if not confirm:
            self.logger.info("Bot-Start abgebrochen durch den Benutzer.")
            return
    
        # Starte den Bot, wenn der Benutzer bestätigt hat
        self.logger.info("Trade-Bot gestartet.")
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
    
        # Starte die Handelslogik in einem separaten Thread
        self.bot_running = True
        threading.Thread(target=self.run_bot_logic, daemon=True).start()
    
    def run_bot_logic(self):
        """Führt die Handelslogik aus, solange der Bot läuft."""
        while self.bot_running:
            try:
                # Warte, bis ein eigenes Order erstellt wurde
                if not hasattr(self, 'own_order') or not self.own_order:
                    self.logger.info("Kein eigenes Order vorhanden. Neues Order wird erstellt.")
                    order_needed, dynamic_price = self.check_if_order_needed()
                    if order_needed:
                        self.execute_bot_trade(dynamic_price)
                    else:
                        self.logger.info("Keine neue Order erforderlich. Warte...")
                        threading.Event().wait(self.check_interval.get())  # Wartezeit aus der GUI
                    continue
    
                # Überprüfen, ob das eigene Order noch existiert
                if not self.check_own_order_exists():
                    self.logger.info("Das eigene Order wurde entfernt. Bot wird gestoppt.")
                    self.stop_bot()
                    break
    
                # Überprüfen, ob eine neue Order erforderlich ist
                order_needed, dynamic_price = self.check_if_order_needed()
                if not order_needed:
                    self.logger.info("Keine neue Order erforderlich. Warte...")
                    threading.Event().wait(self.check_interval.get())  # Wartezeit aus der GUI
                    continue
    
                # Handelslogik ausführen
                self.execute_bot_trade(dynamic_price)
    
            except Exception as e:
                self.logger.error(f"Fehler in der Handelslogik: {str(e)}")
            finally:
                # Warte eine bestimmte Zeit, bevor die nächste Order ausgeführt wird
                threading.Event().wait(self.order_interval.get())  # Wartezeit aus der GUI
    
    def check_own_order_exists(self):
        """Überprüft, ob das eigene Order noch in der Datenbank vorhanden ist."""
        if not hasattr(self, 'own_order') or not self.own_order:
            return False  # Kein eigenes Order vorhanden
    
        try:
            trading_pair = self.own_order.get("trading_pair")
            order_id = self.own_order.get("order_id")
    
            if not trading_pair or not order_id:
                self.logger.error("Ungültige Order-Daten. Überprüfung nicht möglich.")
                self.own_order = None  # Zurücksetzen
                return False
    
            # Hole das Orderbuch für das Handelspaar
            orderbook = self.db_manager.get_orderbook(trading_pair)
    
            # Überprüfen, ob das eigene Order noch in den `bids` oder `asks` vorhanden ist
            for order in orderbook.get('orders', {}).get('bids', []) + orderbook.get('orders', {}).get('asks', []):
                if order[3] == order_id:  # Vergleiche die Order-ID
                    return True  # Order existiert noch
    
            self.logger.info(f"Das eigene Order mit ID {order_id} wurde entfernt.")
            self.own_order = None  # Zurücksetzen, da das Order nicht mehr existiert
            return False
    
        except Exception as e:
            self.logger.error(f"Fehler bei der Überprüfung des eigenen Orders: {str(e)}")
            self.own_order = None  # Zurücksetzen bei Fehlern
            return False

    def check_if_order_needed(self):
        """Prüft, ob eine neue Order erforderlich ist."""
        try:
            # Hole das aktuelle Handelspaar
            selected_display_name = self.selected_pair.get()
            trading_pair = self.pair_mapping.get(selected_display_name)
            orderbook = self.db_manager.get_orderbook(trading_pair)
    
            # Überprüfen, ob das Orderbuch gültige Daten enthält
            if not orderbook or 'orders' not in orderbook:
                self.logger.info("Orderbuch ist leer. Keine Aktion erforderlich.")
                return False, None
    
            # Überprüfen, ob das eigene Order bekannt ist
            own_order = getattr(self, 'own_order', None)
            if not own_order:
                self.logger.info("Kein eigenes Order lokal gespeichert. Abrufen...")
                existing_orders = self.api_client.get_orders()
                for order in existing_orders.get('orders', []):
                    if order['trading_pair'] == trading_pair and order['type'] == self.trade_type.get():
                        self.own_order = order
                        self.logger.info(f"Eigenes Order gefunden: {order['order_id']}")
                        return False, None
                self.logger.info("Kein eigenes Order vorhanden. Neues Order erforderlich.")
                return True, float(self.min_price.get())  # Fallback auf `min_price`
    
            # Eigene Order-Daten
            own_price = float(own_order.get('price', 0))
            own_order_id = own_order.get('order_id')  # ID des eigenen Orders
            trade_type = own_order['type']
            max_price = float(self.max_price.get())
            min_price = float(self.min_price.get())
    
            # Logik für Kauforders (buy)
            if trade_type == "buy":
                bids = orderbook['orders'].get('bids', [])
                if not bids:
                    self.logger.warning("Keine Kauforders ('bids') im Orderbuch verfügbar.")
                    return False, None
    
                # Höchstes Gebot (ohne eigenes Order)
                highest_bid = next((bid for bid in bids if bid[3] != own_order_id), None)
                if not highest_bid:
                    self.logger.info("Kein anderes Gebot im Orderbuch. Keine Aktion erforderlich.")
                    return False, None
    
                highest_bid_price = float(highest_bid[0])
                self.logger.info(f"Höchster Ankaufs-Preis im Orderbuch (ohne eigenes Order): {highest_bid_price}")
    
                # Prüfen, ob das eigene Order das höchste ist
                if own_price > highest_bid_price:
                    # Preis senken, falls möglich
                    new_price = max(highest_bid_price + 0.01, min_price)
                    if new_price != own_price:
                        self.logger.info(f"Preisänderung erforderlich: Aktueller Preis {own_price}, neuer Preis {new_price}.")
                        return True, new_price                
                elif own_price < highest_bid_price:
                    # Preis erhöhen, falls nötig
                    if highest_bid_price > max_price:
                        # Setze auf den nächsthöheren Preis innerhalb der Spanne
                        next_highest_bid = next((bid for bid in bids if bid[3] != own_order_id and float(bid[0]) < max_price), None)
                        if next_highest_bid:
                            next_highest_price = float(next_highest_bid[0]) + 0.01
                            if next_highest_price <= max_price:
                                # Logge den nächstniedrigeren Preis, da er relevant ist
                                self.logger.info(f"Nächstniedriger Preis im Orderbuch: {float(next_highest_bid[0])}")
                                self.logger.info(f"Preisänderung erforderlich: Aktueller Preis {own_price}, neuer Preis {next_highest_price}.")
                                return True, next_highest_price
                        self.logger.info("Kein gültiger Preis innerhalb der Spanne verfügbar. Keine Aktion erforderlich.")
                        return False, None
                    else:
                        new_price = min(highest_bid_price + 0.01, max_price)
                        if new_price != own_price:
                            self.logger.info(f"Preisänderung erforderlich: Aktueller Preis {own_price}, neuer Preis {new_price}.")
                            return True, new_price
                
                self.logger.info("Das eigene Order hat bereits die beste Position. Keine Aktion erforderlich.")
                return False, None
    
            # Logik für Verkaufsorders (sell)
            elif trade_type == "sell":
                asks = orderbook['orders'].get('asks', [])
                if not asks:
                    self.logger.warning("Keine Verkaufsorders ('asks') im Orderbuch verfügbar.")
                    return False, None
            
                # Niedrigstes Angebot (ohne eigenes Order)
                lowest_ask = next((ask for ask in asks if ask[3] != own_order_id), None)
                if not lowest_ask:
                    self.logger.info("Kein anderes Angebot im Orderbuch. Keine Aktion erforderlich.")
                    return False, None
            
                lowest_ask_price = float(lowest_ask[0])
                self.logger.info(f"Niedrigster Verkaufs-Preis im Orderbuch (ohne eigenes Order): {lowest_ask_price}")
            
                # Prüfen, ob das eigene Order das niedrigste ist
                if own_price < lowest_ask_price:
                    # Preis erhöhen, falls nötig
                    new_price = min(lowest_ask_price - 0.01, max_price)
                    if new_price != own_price:
                        self.logger.info(f"Preisänderung erforderlich: Aktueller Preis {own_price}, neuer Preis {new_price}.")
                        return True, new_price
                elif own_price > lowest_ask_price:
                    # Preis senken, falls möglich
                    if lowest_ask_price < min_price:
                        # Setze auf den nächsthöheren Preis innerhalb der Spanne
                        next_lowest_ask = next((ask for ask in asks if ask[3] != own_order_id and float(ask[0]) > min_price), None)
                        if next_lowest_ask:
                            next_lowest_price = float(next_lowest_ask[0]) - 0.01
                            if next_lowest_price >= min_price:
                                # Logge den nächsthöheren Preis, da er relevant ist
                                self.logger.info(f"Nächsthöherer Preis im Orderbuch: {float(next_lowest_ask[0])}")
                                self.logger.info(f"Preisänderung erforderlich: Aktueller Preis {own_price}, neuer Preis {next_lowest_price}.")
                                return True, next_lowest_price
                        self.logger.info("Kein gültiger Preis innerhalb der Spanne verfügbar. Keine Aktion erforderlich.")
                        return False, None
                    else:
                        new_price = max(lowest_ask_price - 0.01, min_price)
                        if new_price != own_price:
                            self.logger.info(f"Preisänderung erforderlich: Aktueller Preis {own_price}, neuer Preis {new_price}.")
                            return True, new_price
            
                self.logger.info("Das eigene Order hat bereits die beste Position. Keine Aktion erforderlich.")
                return False, None
            
            self.logger.info("Keine Aktion erforderlich.")
            return False, None
    
        except Exception as e:
            self.logger.error(f"Fehler bei der Überprüfung, ob eine Order erforderlich ist: {str(e)}")
            return False, None
  
    def execute_bot_trade(self, dynamic_price):
        """Führt eine Order basierend auf den aktuellen Einstellungen aus."""
        if not self.api_client:
            self.logger.error("API-Client nicht konfiguriert.")
            return
    
        try:
            # Hole das aktuelle Handelspaar
            selected_display_name = self.selected_pair.get()
            trading_pair = self.pair_mapping.get(selected_display_name)
    
            if not trading_pair:
                self.logger.error(f"Ungültiges Handelspaar: {selected_display_name}")
                return
    
            self.logger.info(f"Abfrage des Orderbuchs für Handelspaar: {trading_pair}")
            orderbook = self.db_manager.get_orderbook(trading_pair)
    
            # Logge den Inhalt des Orderbuchs
            self.logger.debug(f"Orderbuch-Daten (asks): {orderbook.get('orders', {}).get('asks', [])}")
            self.logger.debug(f"Orderbuch-Daten (bids): {orderbook.get('orders', {}).get('bids', [])}")
    
            if not orderbook or 'orders' not in orderbook or 'asks' not in orderbook['orders'] or 'bids' not in orderbook['orders']:
                self.logger.error("Orderbuch enthält keine gültigen 'asks' oder 'bids'.")
                return
    
            # Sicherstellen, dass `dynamic_price` gesetzt ist
            min_price = float(self.min_price.get())
            if dynamic_price is None:
                self.logger.warning("`dynamic_price` ist None. Fallback auf `min_price`.")
                dynamic_price = min_price
    
            # Überprüfen, ob ein bestehendes Order vorhanden ist
            if hasattr(self, 'own_order') and self.own_order:
                existing_order_id = self.own_order.get("order_id")
                existing_trading_pair = self.own_order.get("trading_pair")
    
                if existing_order_id and existing_trading_pair:
                    # Lösche das bestehende Order
                    self.logger.info(f"Lösche bestehendes Order: {existing_order_id}")
                    delete_response = self.api_client.delete_order(existing_trading_pair, existing_order_id)
                    if delete_response and not delete_response.get('errors'):
                        self.logger.info(f"Bestehendes Order erfolgreich gelöscht: {existing_order_id}")
                        self.own_order = None  # Lösche die gespeicherte Order
                    else:
                        error_message = delete_response.get('errors', ['Unbekannter Fehler'])[0] if delete_response else 'Keine Antwort von der API'
                        self.logger.error(f"Fehler beim Löschen des bestehenden Orders: {error_message}")
                        return  # Abbrechen, wenn das Löschen fehlschlägt
    
            # Erstelle die Order-Parameter
            order_params = {
                "trading_pair": trading_pair.lower(),
                "type": self.trade_type.get(),
                "max_amount_currency_to_trade": float(self.trade_amount.get()),
                "price": dynamic_price,
                "end_datetime": self.get_default_end_datetime(),
                "min_trust_level": self.min_trust_level.get(),
                "seat_of_bank": [self.seat_of_bank.get().split()[0]]
            }
    
            # Zusätzliche Parameter für `sell`-Orders
            if self.trade_type.get() == "sell":
                payment_option = self.payment_option.get()
                if payment_option > 0:
                    order_params["payment_option"] = payment_option
                    if payment_option in [2, 3]:  # SEPA-Only oder Express & SEPA
                        sepa_value = self.sepa_option.get()
                        if sepa_value is not None:
                            order_params["sepa_option"] = sepa_value
    
            # Neues Order erstellen
            self.logger.info(f"Creating order for {trading_pair}: {order_params}")
            response = self.api_client.create_order(**order_params)
            self.logger.debug(f"API-Antwort beim Erstellen der Order: {response}")
            if response and 'order_id' in response and not response.get('errors'):
                self.logger.info(f"Order erfolgreich erstellt: {response['order_id']}")
                self.own_order = {
                    "order_id": response['order_id'],
                    "trading_pair": trading_pair,
                    "type": self.trade_type.get(),
                    "price": dynamic_price
                }
            else:
                error_message = response.get('errors', ['Unbekannter Fehler'])[0] if response else 'Keine Antwort von der API'
                self.logger.error(f"Fehler beim Erstellen der Order: {error_message}")
    
        except Exception as e:
            self.logger.error(f"Fehler beim Ausführen der Order: {str(e)}")

    def stop_bot(self):
        """Stoppt den Trade-Bot."""
        self.logger.info("Trade-Bot wird gestoppt.")
        self.bot_running = False  # Stoppe die Handelslogik
    
        # Lösche die aktive Order, falls vorhanden
        if hasattr(self, 'own_order') and self.own_order:
            try:
                trading_pair = self.own_order.get("trading_pair")
                order_id = self.own_order.get("order_id")
    
                if not trading_pair or not order_id:
                    self.logger.error("Ungültige Order-Daten. Order konnte nicht gelöscht werden.")
                    self.own_order = None  # Zurücksetzen, um den Zustand zu korrigieren
                    return
    
                delete_response = self.api_client.delete_order(trading_pair, order_id)
                if delete_response and not delete_response.get('errors'):
                    self.logger.info(f"Order erfolgreich gelöscht: {order_id}")
                    self.own_order = None  # Lösche die gespeicherte Order
                else:
                    error_message = delete_response.get('errors', ['Unbekannter Fehler'])[0] if delete_response else 'Keine Antwort von der API'
                    self.logger.error(f"Fehler beim Löschen der Order: {error_message}")
                    self.own_order = None  # Zurücksetzen, auch wenn das Löschen fehlschlägt
            except Exception as e:
                self.logger.error(f"Fehler beim Löschen der Order: {str(e)}")
                self.own_order = None  # Zurücksetzen, um den Zustand zu korrigieren
    
        # Aktualisiere die UI
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
    
        # Aktualisiere die UI
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")

from tkinter import ttk, StringVar, IntVar, BooleanVar, messagebox
from tkinter import filedialog
import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import json
import logging
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

class SettingsTab:
    def __init__(self, parent, logger, api_key_var, api_secret_var, api_basic_var, db_config, save_callback):
        self.parent = parent
        self.logger = logger
        self.api_key_var = api_key_var
        self.api_secret_var = api_secret_var
        self.api_basic_var = api_basic_var
        self.db_config = db_config  # Dictionary for MySQL connection info
        self.save_callback = save_callback
        self.setup_ui()

    def show_info_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showinfo(title, message, parent=self.root)
    
    def show_error_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showerror(title, message, parent=self.root)
    
    def show_warning_message(self, title, message):
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        self.root.iconbitmap(default=icon_path)  # Setzen des Symbols
        messagebox.showwarning(title, message, parent=self.root)

    def setup_ui(self):
        # Create main frame
        main_frame = ttk.Frame(self.parent, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
        # API Settings section
        api_frame = ttk.LabelFrame(main_frame, text="API Einstellungen", padding="5")
        api_frame.grid(row=0, column=0, padx=5, pady=5, sticky='ew')
    
        # API Key
        ttk.Label(api_frame, text="API Schlüssel:").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(api_frame, textvariable=self.api_key_var, width=50).grid(row=0, column=1, padx=5, pady=5)
    
        # API Secret
        ttk.Label(api_frame, text="API Secret:").grid(row=1, column=0, padx=5, pady=5)
        ttk.Entry(api_frame, textvariable=self.api_secret_var, width=50, show="*").grid(row=1, column=1, padx=5, pady=5)
    
        # Basic API Key
        ttk.Label(api_frame, text="Basic API Schlüssel:").grid(row=2, column=0, padx=5, pady=5)
        ttk.Entry(api_frame, textvariable=self.api_basic_var, width=50).grid(row=2, column=1, padx=5, pady=5)
    
        # Database Settings section
        db_frame = ttk.LabelFrame(main_frame, text="Datenbank Einstellungen", padding="5")
        db_frame.grid(row=1, column=0, padx=5, pady=5, sticky='ew')
    
        # MySQL Host
        ttk.Label(db_frame, text="MySQL Host:").grid(row=0, column=0, padx=5, pady=5)
        self.host_var = StringVar(value=self.db_config.get('host', ''))
        ttk.Entry(db_frame, textvariable=self.host_var, width=50).grid(row=0, column=1, padx=5, pady=5)
    
        # MySQL User
        ttk.Label(db_frame, text="MySQL Benutzer:").grid(row=1, column=0, padx=5, pady=5)
        self.user_var = StringVar(value=self.db_config.get('user', ''))
        ttk.Entry(db_frame, textvariable=self.user_var, width=50).grid(row=1, column=1, padx=5, pady=5)
    
        # MySQL Password
        ttk.Label(db_frame, text="MySQL Passwort:").grid(row=2, column=0, padx=5, pady=5)
        self.password_var = StringVar(value=self.db_config.get('password', ''))
        ttk.Entry(db_frame, textvariable=self.password_var, width=50, show="*").grid(row=2, column=1, padx=5, pady=5)
    
        # MySQL Database
        ttk.Label(db_frame, text="MySQL Datenbank:").grid(row=3, column=0, padx=5, pady=5)
        self.database_var = StringVar(value=self.db_config.get('database', ''))
        ttk.Entry(db_frame, textvariable=self.database_var, width=50).grid(row=3, column=1, padx=5, pady=5)
    
        # MySQL Port
        ttk.Label(db_frame, text="MySQL Port:").grid(row=4, column=0, padx=5, pady=5)
        self.port_var = IntVar(value=self.db_config.get('port', 3306))  # Default to 3306
        ttk.Entry(db_frame, textvariable=self.port_var, width=50).grid(row=4, column=1, padx=5, pady=5)
    
        # Save Button
        ttk.Button(main_frame, text="Einstellungen Speichern", command=self.save_settings).grid(row=2, column=0, pady=10)
    
        # Add HTTP Response Codes Button
        ttk.Button(main_frame, text="Definition Error Codes", command=self.show_http_response_codes).grid(row=3, column=0, pady=5)

        # Add Console section
        console_frame = ttk.LabelFrame(main_frame, text="Console Ausgabe", padding="5")
        console_frame.grid(row=4, column=0, padx=5, pady=5, sticky='nsew')
    
        # Create console text widget with scrollbar
        self.console = tk.Text(console_frame, height=15, width=140)
        scrollbar = ttk.Scrollbar(console_frame, orient='vertical', command=self.console.yview)
        self.console.configure(yscrollcommand=scrollbar.set)
    
        # Grid layout for console and scrollbar
        self.console.grid(row=0, column=0, padx=5, pady=5, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
    
        # Configure grid weights for expandable console
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)
        console_frame.columnconfigure(0, weight=1)
        console_frame.rowconfigure(0, weight=1)
    
        # Add console handler to logger
        self.add_console_handler()

    def save_settings(self):
        """Save the settings to the credentials manager"""
        try:
            self.db_config['host'] = self.host_var.get()
            self.db_config['user'] = self.user_var.get()
            self.db_config['password'] = self.password_var.get()
            self.db_config['database'] = self.database_var.get()
            self.db_config['port'] = self.port_var.get()
            
            # Call the save callback with the updated db_config
            self.save_callback(self.db_config)
            
            messagebox.showinfo("Erfolg", "Einstellungen erfolgreich gespeichert")
        except Exception as e:
            self.logger.error(f"Fehler beim Speichern der Einstellungen: {str(e)}")
            messagebox.showerror("Fehler", f"Fehler beim Speichern der Einstellungen: {str(e)}")

    def add_console_handler(self):
        class ConsoleHandler(logging.Handler):
            def __init__(self, console):
                logging.Handler.__init__(self)
                self.console = console

            def emit(self, record):
                msg = self.format(record) + '\n'
                self.console.insert('end', msg)
                self.console.see('end')

        console_handler = ConsoleHandler(self.console)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.addHandler(console_handler)
    
    def show_http_response_codes(self):
        """Show a window with HTTP response codes"""
        response_codes = """
        General HTTP Response Codes
        Code    Message
        200     GET-/ DELETE-Request wurde erfolgreich durchgeführt
        201     POST-Request wurde erfolgreich durchgeführt und die neue Ressource angelegt (z.B. Trade)
        400     Bad Request
        403     Forbidden
        404     Angefragte Entität konnte nicht gefunden werden
        422     Bitte die im Response aufgeführten Fehler im Array-Eintrag "errors" untersuchen.
        429     Too many requests

        Request
        Error-Code  Message

        1   Missing header
        2   Inactive api key
        3   Invalid api key
        4   Invalid nonce
        5   Invalid signature
        6   Insufficient credits
        7   Invalid route
        8   Unkown api action
        9   Additional agreement not accepted
        32  Api key banned
        33  Ip banned
        94  Ip access restricted
        10  No 2 factor authentication
        11  No beta group user
        12  Technical reason
        13  Trading api currently unavailable
        14  No action permission for api key
        15  Missing post parameter
        16  Missing get parameter
        17  Invalid number
        18  Number too low
        19  Number too big
        20  Too many decimal places
        21  Invalid boolean value
        22  Forbidden parameter value
        23  Invalid min amount
        24  Invalid datetime format
        25  Date lower than min date
        26  Invalid value
        27  Forbidden value for get parameter
        28  Forbidden value for post parameter
        29  Express trade temporarily not available
        30  End datetime younger than start datetime
        31  Page greater than last page
        34  Invalid trading pair
        35  Invalid currency
        36  Forbidden value for query parameter
        37  Too many characters
        44  No kyc full
        45  Operation currently not possible
        46  Has to accept additional agreement
        47  Not part of beta group for api version
        113 Futurum not possible
        114 Futurum outside business hours
        116 Trading pair is delisted
        117 Currency is delisted

        Order
        Error-Code  Message
        50  Order not found
        51  Order not possible
        52  Invalid order type
        53  Payment option not allowed for type buy
        54  Cancellation not allowed
        55  Trading suspended
        56  Express trade not possible
        106 Sepa trade not possible
        57  No bank account
        58  Order not possible for trading pair
        59  Order not possible due supplementary agreement
        107 Sepa instant not allowed for type buy
        108 Sepa instant not allowed for payment option
        109 Sepa instant order not possible

        Trade
        Error-Code  Message
        70  No active reservation
        71  Express trade not allowed
        72  Express trade failure temporary
        73  Express trade failure
        74  Invalid trade state
        75  Trade not found
        76  Reservation amount insufficient
        77  Volume currency to pay after fee deviates
        78  Already marked as paid
        79  Payment already marked as received
        100 Trade already rated
        101 Trade not possible for trading pair
        103 Already marked as transferred
        104 Coins already marked as received
        110 Sepa instant trade not possible
        115 Confirmation of payment receipt currently not possible
        """
        # Create a new window
        response_window = tk.Toplevel(self.parent)
        response_window.title("HTTP Response Codes")
        icon_path = os.path.join(os.path.dirname(__file__), 'bitcoin.ico')
        response_window.iconbitmap(default=icon_path)  # Setzen des Symbols

        # Create a text widget with larger font
        text_widget = tk.Text(response_window, wrap='word', font=("Helvetica", 12))
        text_widget.insert('1.0', response_codes)
        text_widget.config(state='disabled')  # Make the text widget read-only

        # Add a scrollbar to the text widget
        scrollbar = ttk.Scrollbar(response_window, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        # Grid layout for text widget and scrollbar
        text_widget.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        # Configure grid weights for expandable text widget
        response_window.columnconfigure(0, weight=1)
        response_window.rowconfigure(0, weight=1)